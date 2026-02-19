"""
Telegram Bot MVP - Invoice Extraction
Handles photo/PDF uploads with confirmation before processing.
"""
import asyncio
import io
import base64
from datetime import datetime
from typing import Dict, Optional
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ContextTypes,
    filters,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from config import settings
from app.services.gemini_service import gemini_service
from app.services.supabase_service import supabase_service
from app.services.storage_service import storage_service
from app.schemas.invoice_schema import ProcessingResponse, InvoiceSchema
from app.utils.logger import setup_logger

logger = setup_logger("telegram_bot")

# Temporary storage for pending files (file_id -> file_info)
pending_files: Dict[str, Dict] = {}

# Temporary storage for pending validations (validation_key -> extracted_data)
pending_validations: Dict[str, Dict] = {}


def create_excel_from_invoice(invoice: InvoiceSchema) -> io.BytesIO:
    """Create Excel file from a single invoice"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Factura"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Invoice summary
    ws.cell(row=1, column=1, value="Campo").fill = header_fill
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=2, value="Valor").fill = header_fill
    ws.cell(row=1, column=2).font = header_font
    
    row = 2
    ws.cell(row=row, column=1, value="Nro. Factura")
    ws.cell(row=row, column=2, value=invoice.invoice_number or "")
    row += 1
    
    ws.cell(row=row, column=1, value="Fecha")
    ws.cell(row=row, column=2, value=invoice.invoice_date or "")
    row += 1
    
    ws.cell(row=row, column=1, value="Proveedor")
    ws.cell(row=row, column=2, value=invoice.supplier_name or "")
    row += 1
    
    ws.cell(row=row, column=1, value="RUC Proveedor")
    ws.cell(row=row, column=2, value=invoice.supplier_ruc or "")
    row += 1
    
    ws.cell(row=row, column=1, value="Cliente")
    ws.cell(row=row, column=2, value=invoice.customer_name or "")
    row += 1
    
    ws.cell(row=row, column=1, value="Subtotal")
    ws.cell(row=row, column=2, value=invoice.subtotal)
    row += 1
    
    ws.cell(row=row, column=1, value="IGV")
    ws.cell(row=row, column=2, value=invoice.tax)
    row += 1
    
    ws.cell(row=row, column=1, value="Total")
    ws.cell(row=row, column=2, value=invoice.total)
    ws.cell(row=row, column=2).font = Font(bold=True)
    row += 1
    
    ws.cell(row=row, column=1, value="Moneda")
    ws.cell(row=row, column=2, value=invoice.currency or "PEN")
    row += 1
    
    # Items section
    if invoice.items:
        row += 2
        ws.cell(row=row, column=1, value="Items").font = Font(bold=True, size=12)
        row += 1
        
        # Item headers
        item_headers = ["#", "Descripción", "Cantidad", "Precio Unit.", "Total"]
        for col, header in enumerate(item_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        row += 1
        
        # Items
        for idx, item in enumerate(invoice.items, 1):
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=item.description or "")
            ws.cell(row=row, column=3, value=item.quantity)
            ws.cell(row=row, column=4, value=item.unit_price)
            ws.cell(row=row, column=5, value=item.total_price)
            row += 1
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file


async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle /start command"""
    chat_id = str(update.effective_chat.id)
    chat_title = update.effective_chat.title or update.effective_user.first_name or "Usuario"
    
    # Register company in Supabase if not exists
    if supabase_service.is_enabled():
        company = await supabase_service.get_company(chat_id)
        if not company:
            await supabase_service.save_company(
                chat_id=chat_id,
                name=chat_title,
                plan="free",
                limit_monthly=100,
                registered_by=update.effective_user.username or str(update.effective_user.id)
            )
            logger.log_info(
                "New company registered",
                chat_id=chat_id,
                name=chat_title
            )
    
    welcome_message = f"""
¡Bienvenido! 👋

Soy un bot de extracción de facturas. 

📸 **Envíame una foto o PDF de tu factura** y te enviaré un Excel con todos los datos extraídos.

📊 **Plan actual**: Gratuito (100 facturas/mes)

💡 **Instrucciones**:
1. Envía foto o PDF de la factura
2. Confirma el procesamiento
3. Recibe tu Excel con los datos

¡Empecemos!
"""
    
    await update.message.reply_text(welcome_message)
    
    logger.log_info(
        "Bot started",
        chat_id=chat_id,
        chat_title=chat_title,
        user=update.effective_user.username
    )


async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle photo uploads"""
    try:
        chat_id = str(update.effective_chat.id)
        # Get largest photo
        photo = update.message.photo[-1]
        file_id = photo.file_id
        file_unique_id = photo.file_unique_id
        
        # Use shorter callback_data to avoid issues
        callback_key = f"{chat_id}_{update.message.message_id}"
        
        # Store pending file
        pending_files[callback_key] = {
            "type": "photo",
            "file_id": file_id,
            "file_unique_id": file_unique_id,
            "chat_id": chat_id,
            "message_id": update.message.message_id,
            "user": update.effective_user.username or str(update.effective_user.id)
        }
        
        logger.log_info(
            "Photo stored in pending_files",
            callback_key=callback_key,
            total_pending=len(pending_files),
            pending_keys=list(pending_files.keys())
        )

        # Create confirmation buttons
        keyboard = [
            [
                InlineKeyboardButton("✅ Sí", callback_data=f"yes_{callback_key}"),
                InlineKeyboardButton("❌ No", callback_data=f"no_{callback_key}")
            ]
        ]
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await update.message.reply_text(
            "📸 Foto recibida.\n\n¿Procesar esta factura?",
            reply_markup=reply_markup
        )
        
        logger.log_info(
            "Photo received, awaiting confirmation",
            chat_id=chat_id,
            file_id=file_id
        )
        
    except Exception as e:
        logger.log_error(
            "Error handling photo",
            error=str(e),
            chat_id=str(update.effective_chat.id) if update.effective_chat else "unknown"
        )
        await update.message.reply_text(
            "❌ Error al recibir la foto. Por favor, intenta nuevamente."
        )


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle document uploads (PDFs)"""
    try:
        chat_id = str(update.effective_chat.id)
        document = update.message.document
        
        file_id = document.file_id
        file_unique_id = document.file_unique_id
        file_name = document.file_name
        mime_type = document.mime_type
        
        # Check if it's a PDF
        if mime_type != "application/pdf":
            await update.message.reply_text(
                "⚠️ Solo acepto archivos PDF o fotos.\n\nPor favor, envía una imagen o PDF de la factura."
            )
            return
        
        # Use shorter callback_data
        callback_key = f"{chat_id}_{update.message.message_id}"
        
        # Store pending file
        pending_files[callback_key] = {
            "type": "pdf",
            "file_id": file_id,
            "file_unique_id": file_unique_id,
            "file_name": file_name,
            "chat_id": chat_id,
            "message_id": update.message.message_id,
            "user": update.effective_user.username or str(update.effective_user.id)
        }
        
        logger.log_info(
            "PDF stored in pending_files",
            callback_key=callback_key,
            file_name=file_name,
            total_pending=len(pending_files),
            pending_keys=list(pending_files.keys())
        )
        
        # Create confirmation buttons
        keyboard = [
            [
                InlineKeyboardButton("✅ Sí", callback_data=f"yes_{callback_key}"),
                InlineKeyboardButton("❌ No", callback_data=f"no_{callback_key}")
            ]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await update.message.reply_text(
            f"📄 PDF recibido: {file_name}\n\n¿Procesar esta factura?",
            reply_markup=reply_markup
        )
        
        logger.log_info(
            "PDF received, awaiting confirmation",
            chat_id=chat_id,
            file_id=file_id,
            file_name=file_name
        )
        
    except Exception as e:
        logger.log_error(
            "Error handling document",
            error=str(e),
            chat_id=str(update.effective_chat.id) if update.effective_chat else "unknown"
        )
        await update.message.reply_text(
            "❌ Error al recibir el documento. Por favor, intenta nuevamente."
        )


async def button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle button callbacks"""
    query = update.callback_query
    await query.answer()
    
    callback_data = query.data
    
    # Log callback received
    logger.log_info(
        "Callback received",
        callback_data=callback_data,
        pending_files_keys=list(pending_files.keys())
    )
    
    # Parse callback: "yes_chatid_msgid" or "no_chatid_msgid"
    parts = callback_data.split("_", 1)
    if len(parts) != 2:
        await query.edit_message_text("⚠️ Solicitud inválida. Por favor, envía la factura nuevamente.")
        return
    
    action, callback_key = parts
    
    # Log parsed data
    logger.log_info(
        "Callback parsed",
        action=action,
        callback_key=callback_key,
        exists=callback_key in pending_files
    )
    
    # Get pending file info
    file_info = pending_files.get(callback_key)
    if not file_info:
        await query.edit_message_text(
            f"⚠️ Archivo expirado o bot reiniciado.\n\n"
            f"Por favor, envía la factura nuevamente.\n\n"
            f"💡 Tip: Procesa la factura inmediatamente después de enviarla."
        )
        logger.log_warning(
            "Pending file not found",
            callback_key=callback_key,
            available_keys=list(pending_files.keys())
        )
        return
    
    if action == "no":
        # Remove from pending
        del pending_files[callback_key]
        await query.edit_message_text("❌ Procesamiento cancelado.")
        logger.log_info("Processing cancelled", callback_key=callback_key)
        return
    
    if action == "yes":
        await query.edit_message_text("⏳ Procesando factura, por favor espera...")
        
        try:
            # Download file from Telegram
            file_id = file_info["file_id"]
            file = await context.bot.get_file(file_id)
            file_bytes = await file.download_as_bytearray()
            
            logger.log_info(
                "File downloaded from Telegram",
                file_id=file_id,
                size_bytes=len(file_bytes)
            )
            
            # Process based on file type
            if file_info["type"] == "photo":
                # Convert bytes to base64
                image_base64 = base64.b64encode(bytes(file_bytes)).decode('utf-8')
                
                # Process as image
                result = await gemini_service.extract_invoice_from_image(
                    image_base64=image_base64,
                    filename=f"telegram_photo_{file_id}.jpg",
                    sequence_id=1
                )
            else:
                # Process as PDF
                logger.log_info("Processing PDF file", callback_key=callback_key)
                
                # Download PDF file
                file = await context.bot.get_file(file_id)
                file_bytes = await file.download_as_bytearray()
                
                # Convert to base64
                pdf_base64 = base64.b64encode(bytes(file_bytes)).decode('utf-8')
                
                # Process with Gemini
                result = await gemini_service.extract_invoice_from_pdf(
                    pdf_base64=pdf_base64,
                    filename=file_info.get("file_name", f"telegram_pdf_{file_id}.pdf"),
                    sequence_id=1
                )
            
            # Check if extraction was successful
            if not result or not result.get("success"):
                error_msg = result.get("error", "Error desconocido") if result else "No se pudo procesar"
                await query.edit_message_text(
                    f"❌ No pude extraer datos de esta factura.\n\n"
                    f"Error: {error_msg}\n\n"
                    f"Intenta con:\n"
                    f"• Una imagen más clara\n"
                    f"• Mejor iluminación\n"
                    f"• Asegúrate de que la factura sea visible"
                )
                del pending_files[callback_key]
                return
            
            # Convert dict to InvoiceSchema object
            invoice_data = result.get("invoice_data", {})
            invoice = InvoiceSchema(**invoice_data)
            
            logger.log_info(
                "Invoice extracted successfully",
                callback_key=callback_key,
                invoice_number=invoice.invoice_number,
                total=invoice.total
            )
            
            # Store extracted data for validation
            validation_key = f"val_{file_info['chat_id']}_{datetime.now().timestamp()}"
            pending_validations[validation_key] = {
                "invoice": invoice,
                "file_info": file_info,
                "file_id": file_id
            }
            
            # Show extracted data for validation
            items_summary = ""
            if invoice.items and len(invoice.items) > 0:
                items_summary = f"\n📦 **Items:** {len(invoice.items)} producto(s)"
                for idx, item in enumerate(invoice.items[:3], 1):
                    items_summary += f"\n   {idx}. {item.description or 'Sin descripción'}" 
                    items_summary += f" - {item.quantity}x S/ {item.unit_price or 0:.2f}"
                if len(invoice.items) > 3:
                    items_summary += f"\n   ... y {len(invoice.items) - 3} más"
            
            validation_message = (
                f"📋 **Datos extraídos por IA**\n\n"
                f"📄 **Nro. Factura:** {invoice.invoice_number or 'N/A'}\n"
                f"📅 **Fecha:** {invoice.invoice_date or 'N/A'}\n"
                f"🏪 **Proveedor:** {invoice.supplier_name or 'N/A'}\n"
                f"🆔 **RUC:** {invoice.supplier_ruc or 'N/A'}\n"
                f"👤 **Cliente:** {invoice.customer_name or 'N/A'}\n\n"
                f"💰 **Subtotal:** {invoice.currency or 'PEN'} {invoice.subtotal or 0:.2f}\n"
                f"📊 **IGV:** {invoice.currency or 'PEN'} {invoice.tax or 0:.2f}\n"
                f"💵 **Total:** {invoice.currency or 'PEN'} {invoice.total or 0:.2f}\n"
                f"{items_summary}\n\n"
                f"✅ **¿Los datos son correctos?**"
            )
            
            # Create validation buttons
            keyboard = [
                [
                    InlineKeyboardButton("✅ Guardar", callback_data=f"save_{validation_key}"),
                    InlineKeyboardButton("✏️ Editar", callback_data=f"edit_{validation_key}")
                ],
                [InlineKeyboardButton("❌ Cancelar", callback_data=f"cancel_{validation_key}")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await query.edit_message_text(validation_message, reply_markup=reply_markup)
            
            # Remove from pending files
            del pending_files[callback_key]
            
            logger.log_info(
                "Processing completed successfully",
                callback_key=callback_key,
                chat_id=file_info["chat_id"]
            )
            
        except Exception as e:
            logger.log_error(
                "Error processing file",
                callback_key=callback_key,
                error=str(e)
            )
            
            await query.edit_message_text(
                f"❌ Error al procesar la factura.\n\n"
                f"Error: {str(e)}\n\n"
                f"Por favor, intenta nuevamente."
            )
            
            # Remove from pending
            if callback_key in pending_files:
                del pending_files[callback_key]
    
    # Handle validation callbacks
    elif callback_data.startswith("save_"):
        validation_key = callback_data.replace("save_", "")
        validation_data = pending_validations.get(validation_key)
        
        if not validation_data:
            await query.edit_message_text("⚠️ Sesión expirada. Por favor, procesa la factura nuevamente.")
            return
        
        await query.edit_message_text("💾 Guardando factura...")
        
        try:
            invoice = validation_data["invoice"]
            file_info = validation_data["file_info"]
            file_id = validation_data["file_id"]
            
            # Generate Excel
            excel_file = create_excel_from_invoice(invoice)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_filename = f"factura_{timestamp}.xlsx"
            
            # Upload to Cloud Storage (optional)
            excel_blob_path = None
            if storage_service.client:
                try:
                    excel_metadata = storage_service.upload_excel(
                        excel_file,
                        excel_filename,
                        company_id=file_info["chat_id"]
                    )
                    if excel_metadata:
                        excel_blob_path = excel_metadata.get("blob_path")
                    excel_file.seek(0)
                except Exception as storage_error:
                    logger.log_warning(
                        "Failed to upload to Cloud Storage",
                        error=str(storage_error)
                    )
            
            # Save to Supabase
            record_id = None
            if supabase_service.is_enabled():
                try:
                    record_id = await supabase_service.save_processing_record(
                        chat_id=file_info["chat_id"],
                        zip_blob_path=None,
                        excel_blob_path=excel_blob_path,
                        total_files=1,
                        success_files=1,
                        error_files=0,
                        telegram_file_id=file_id,
                        telegram_file_unique_id=file_info["file_unique_id"]
                    )
                    
                    if record_id:
                        await supabase_service.save_invoices_batch(
                            invoices=[invoice],
                            chat_id=file_info["chat_id"],
                            record_id=record_id
                        )
                    
                    logger.log_info(
                        "Data saved to Supabase",
                        record_id=record_id,
                        chat_id=file_info["chat_id"]
                    )
                except Exception as db_error:
                    logger.log_error(
                        "Failed to save to Supabase",
                        error=str(db_error)
                    )
            
            # Send Excel to user
            caption = (
                f"✅ Factura guardada exitosamente\n\n"
                f"💰 Total: {invoice.currency or 'PEN'} {invoice.total or 0:.2f}\n"
                f"📊 IGV: {invoice.currency or 'PEN'} {invoice.tax or 0:.2f}\n"
                f"📄 Nro: {invoice.invoice_number or 'N/A'}"
            )
            
            await context.bot.send_document(
                chat_id=file_info["chat_id"],
                document=excel_file,
                filename=excel_filename,
                caption=caption
            )
            
            await query.edit_message_text("✅ ¡Factura guardada y Excel enviado!")
            
            # Remove from pending validations
            del pending_validations[validation_key]
            
        except Exception as e:
            logger.log_error("Error saving validated invoice", error=str(e))
            await query.edit_message_text(f"❌ Error al guardar: {str(e)}")
    
    elif callback_data.startswith("edit_"):
        validation_key = callback_data.replace("edit_", "")
        validation_data = pending_validations.get(validation_key)
        
        if not validation_data:
            await query.edit_message_text("⚠️ Sesión expirada. Por favor, procesa la factura nuevamente.")
            return
        
        # Show editable fields
        keyboard = [
            [InlineKeyboardButton("📄 Nro. Factura", callback_data=f"field_invoice_number_{validation_key}")],
            [InlineKeyboardButton("📅 Fecha", callback_data=f"field_invoice_date_{validation_key}")],
            [InlineKeyboardButton("🏪 Proveedor", callback_data=f"field_supplier_name_{validation_key}")],
            [InlineKeyboardButton("🆔 RUC Proveedor", callback_data=f"field_supplier_ruc_{validation_key}")],
            [InlineKeyboardButton("👤 Cliente", callback_data=f"field_customer_name_{validation_key}")],
            [InlineKeyboardButton("💰 Subtotal", callback_data=f"field_subtotal_{validation_key}")],
            [InlineKeyboardButton("📊 IGV", callback_data=f"field_tax_{validation_key}")],
            [InlineKeyboardButton("💵 Total", callback_data=f"field_total_{validation_key}")],
            [InlineKeyboardButton("🔙 Volver", callback_data=f"back_{validation_key}")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(
            "✏️ **Selecciona el campo a editar:**",
            reply_markup=reply_markup
        )
    
    elif callback_data.startswith("field_"):
        # Parse: "field_invoice_number_val_123" -> field_name="invoice_number", validation_key="val_123"
        # Remove "field_" prefix
        remaining = callback_data.replace("field_", "", 1)
        
        # Find where validation_key starts (it always starts with "val_")
        val_index = remaining.find("_val_")
        if val_index != -1:
            field_name = remaining[:val_index]
            validation_key = remaining[val_index + 1:]  # Skip the leading underscore
            
            # Store the field being edited in context
            context.user_data['editing_field'] = field_name
            context.user_data['editing_validation_key'] = validation_key
            
            field_labels = {
                "invoice_number": "número de factura",
                "invoice_date": "fecha (DD-MM-YYYY)",
                "supplier_name": "nombre del proveedor",
                "supplier_ruc": "RUC del proveedor",
                "customer_name": "nombre del cliente",
                "subtotal": "subtotal",
                "tax": "IGV",
                "total": "total"
            }
            
            field_label = field_labels.get(field_name, field_name)
            
            await query.edit_message_text(
                f"✏️ Envía el nuevo valor para **{field_label}**\n\n"
                f"O usa /cancelar para volver"
            )
    
    elif callback_data.startswith("back_"):
        validation_key = callback_data.replace("back_", "")
        validation_data = pending_validations.get(validation_key)
        
        if not validation_data:
            await query.edit_message_text("⚠️ Sesión expirada. Por favor, procesa la factura nuevamente.")
            return
        
        # Show validation summary again
        invoice = validation_data["invoice"]
        
        items_summary = ""
        if invoice.items and len(invoice.items) > 0:
            items_summary = f"\n📦 **Items:** {len(invoice.items)} producto(s)"
            for idx, item in enumerate(invoice.items[:3], 1):
                items_summary += f"\n   {idx}. {item.description or 'Sin descripción'}"
                items_summary += f" - {item.quantity}x S/ {item.unit_price or 0:.2f}"
            if len(invoice.items) > 3:
                items_summary += f"\n   ... y {len(invoice.items) - 3} más"
        
        validation_message = (
            f"📋 **Datos extraídos por IA**\n\n"
            f"📄 **Nro. Factura:** {invoice.invoice_number or 'N/A'}\n"
            f"📅 **Fecha:** {invoice.invoice_date or 'N/A'}\n"
            f"🏪 **Proveedor:** {invoice.supplier_name or 'N/A'}\n"
            f"🆔 **RUC:** {invoice.supplier_ruc or 'N/A'}\n"
            f"👤 **Cliente:** {invoice.customer_name or 'N/A'}\n\n"
            f"💰 **Subtotal:** {invoice.currency or 'PEN'} {invoice.subtotal or 0:.2f}\n"
            f"📊 **IGV:** {invoice.currency or 'PEN'} {invoice.tax or 0:.2f}\n"
            f"💵 **Total:** {invoice.currency or 'PEN'} {invoice.total or 0:.2f}\n"
            f"{items_summary}\n\n"
            f"✅ **¿Los datos son correctos?**"
        )
        
        keyboard = [
            [
                InlineKeyboardButton("✅ Guardar", callback_data=f"save_{validation_key}"),
                InlineKeyboardButton("✏️ Editar", callback_data=f"edit_{validation_key}")
            ],
            [InlineKeyboardButton("❌ Cancelar", callback_data=f"cancel_{validation_key}")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(validation_message, reply_markup=reply_markup)
    
    elif callback_data.startswith("cancel_"):
        validation_key = callback_data.replace("cancel_", "")
        
        if validation_key in pending_validations:
            del pending_validations[validation_key]
        
        await query.edit_message_text("❌ Procesamiento cancelado")
    
    # Handle delete callbacks
    elif callback_data.startswith("delete_"):
        chat_id = str(update.effective_chat.id)
        
        if callback_data == "delete_cancel":
            await query.edit_message_text("❌ Operación cancelada")
            return
        
        # Extract invoice_id
        invoice_id = callback_data.replace("delete_", "")
        
        # Confirm deletion
        keyboard = [
            [
                InlineKeyboardButton("⚠️ Sí, eliminar", callback_data=f"confirm_delete_{invoice_id}"),
                InlineKeyboardButton("❌ No, cancelar", callback_data="delete_cancel")
            ]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(
            "⚠️ **¿Estás seguro?**\n\n"
            "Esta factura será **eliminada permanentemente** de la base de datos.\n"
            "❗ Esta acción NO se puede deshacer.",
            reply_markup=reply_markup
        )
    
    elif callback_data.startswith("confirm_delete_"):
        chat_id = str(update.effective_chat.id)
        invoice_id = callback_data.replace("confirm_delete_", "")
        
        try:
            if not supabase_service.is_enabled():
                await query.edit_message_text("❌ Servicio de base de datos no disponible.")
                return
            
            # Delete invoice items first (due to foreign key)
            supabase_service.client.table("invoice_items").delete().eq(
                "invoice_id", invoice_id
            ).eq("company_id", chat_id).execute()
            
            # Delete invoice (hard delete)
            result = supabase_service.client.table("invoices").delete().eq(
                "id", invoice_id
            ).eq("company_id", chat_id).execute()
            
            if result.data:
                await query.edit_message_text("✅ Factura eliminada permanentemente de la base de datos")
                logger.log_info("Invoice permanently deleted", chat_id=chat_id, invoice_id=invoice_id)
            else:
                await query.edit_message_text("❌ No se pudo eliminar la factura")
        
        except Exception as e:
            logger.log_error("Error deleting invoice", error=str(e), chat_id=chat_id)
            await query.edit_message_text(f"❌ Error: {str(e)}")


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle /ayuda command"""
    help_text = """
📖 **AYUDA - Bot de Facturas para PYMEs**

**📸 PROCESAR FACTURAS:**
1. Envía una foto clara de tu factura o un PDF
2. Confirma con ✅ Sí
3. Recibe Excel con los datos extraídos

**Formatos aceptados:**
• 📸 Fotos (JPG, PNG)
• 📄 PDFs

**💰 CONSULTAS DE GASTOS:**

`/resumen DD-MM-YYYY DD-MM-YYYY`
Resumen de gastos en un periodo
Ejemplo: `/resumen 01-01-2026 31-01-2026`
📊 Total gastado, cantidad de facturas, Excel del periodo

`/proveedores [mes]` o `/proveedores [inicio] [fin]`
Top 10 proveedores por gasto
Ejemplo: `/proveedores 01-2026`
🏪 Dónde gastas más, útil para negociar

`/estadisticas [mes]`
Dashboard de estadísticas del mes
Ejemplo: `/estadisticas 01-2026`
📈 Promedio, día con más gastos, top proveedores

`/comparar [mes1] [mes2]`
Compara gastos entre dos meses
Ejemplo: `/comparar 01-2026 12-2025`
📊 % de aumento/reducción de gastos

**🔍 BUSCAR FACTURAS:**

`/buscar [término]`
Buscar por número, proveedor, RUC o monto
Ejemplo: `/buscar F001-12345` o `/buscar Sodimac`

`/items [producto]` o `/items [mes]`
Buscar productos y ver precios históricos
Ejemplo: `/items laptop` o `/items 01-2026`
💻 Detecta si te cobran más caro

`/historial [cantidad]`
Ver últimas N facturas procesadas
Ejemplo: `/historial 10`
📄 Con botones para descargar Excel

**🗑️ GESTIÓN:**

`/eliminar`
Ver últimas facturas y eliminar errores
Pide confirmación antes de borrar

**⚙️ OTROS:**

`/start` - Registrar tu empresa
`/ayuda` - Ver esta ayuda

**💡 Consejos:**
• Foto clara, buena iluminación
• Factura completa visible
• PDFs directamente desde el emisor
• Evita sombras y reflejos en fotos
"""
    
    await update.message.reply_text(help_text)


async def resumen_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle /resumen [fecha_inicio] [fecha_fin]"""
    chat_id = str(update.effective_chat.id)
    
    try:
        # Parse dates from command
        if len(context.args) < 2:
            await update.message.reply_text(
                "❌ Formato incorrecto.\n\n"
                "Uso: `/resumen DD-MM-YYYY DD-MM-YYYY`\n"
                "Ejemplo: `/resumen 01-01-2026 31-01-2026`"
            )
            return
        
        fecha_inicio_str = context.args[0]
        fecha_fin_str = context.args[1]
        
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, "%d-%m-%Y").date()
            fecha_fin = datetime.strptime(fecha_fin_str, "%d-%m-%Y").date()
        except ValueError:
            await update.message.reply_text(
                "❌ Formato de fecha inválido.\n\n"
                "Usa: DD-MM-YYYY\n"
                "Ejemplo: `/resumen 01-01-2026 31-01-2026`"
            )
            return
        
        if fecha_inicio > fecha_fin:
            await update.message.reply_text("❌ La fecha de inicio debe ser anterior a la fecha de fin.")
            return
        
        # Get invoices from Supabase
        if not supabase_service.is_enabled():
            await update.message.reply_text("❌ Servicio de base de datos no disponible.")
            return
        
        await update.message.reply_text("⏳ Consultando facturas...")
        
        # Query invoices in date range
        result = supabase_service.client.table("invoices").select("*").eq(
            "company_id", chat_id
        ).gte("invoice_date", str(fecha_inicio)).lte("invoice_date", str(fecha_fin)).is_(
            "deleted_at", "null"
        ).execute()
        
        invoices = result.data if result.data else []
        
        if not invoices:
            await update.message.reply_text(
                f"📭 No hay facturas entre {fecha_inicio_str} y {fecha_fin_str}"
            )
            return
        
        # Calculate totals
        total_subtotal = sum(float(inv.get("subtotal", 0) or 0) for inv in invoices)
        total_igv = sum(float(inv.get("igv", 0) or 0) for inv in invoices)
        total_general = sum(float(inv.get("total", 0) or 0) for inv in invoices)
        cantidad = len(invoices)
        promedio = total_general / cantidad if cantidad > 0 else 0
        currency = invoices[0].get("currency", "PEN") if invoices else "PEN"
        
        # Generate Excel with all invoices
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen"
        
        # Header
        headers = ["Fecha", "Nro. Factura", "Proveedor", "RUC", "Subtotal", "IGV", "Total", "Moneda"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Data rows
        for row_idx, inv in enumerate(invoices, 2):
            ws.cell(row=row_idx, column=1, value=inv.get("invoice_date", ""))
            ws.cell(row=row_idx, column=2, value=inv.get("invoice_number", ""))
            ws.cell(row=row_idx, column=3, value=inv.get("supplier_name", ""))
            ws.cell(row=row_idx, column=4, value=inv.get("supplier_ruc", ""))
            ws.cell(row=row_idx, column=5, value=float(inv.get("subtotal", 0) or 0))
            ws.cell(row=row_idx, column=6, value=float(inv.get("igv", 0) or 0))
            ws.cell(row=row_idx, column=7, value=float(inv.get("total", 0) or 0))
            ws.cell(row=row_idx, column=8, value=inv.get("currency", "PEN"))
        
        # Totals row
        total_row = len(invoices) + 2
        ws.cell(row=total_row, column=4, value="TOTAL:").font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=total_subtotal).font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=total_igv).font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=total_general).font = Font(bold=True)
        
        # Auto-adjust columns
        for col in range(1, 9):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Send summary message
        summary = (
            f"📊 **Resumen de Gastos**\n"
            f"📅 Periodo: {fecha_inicio_str} - {fecha_fin_str}\n\n"
            f"💰 **Total:** {currency} {total_general:,.2f}\n"
            f"📄 Subtotal: {currency} {total_subtotal:,.2f}\n"
            f"💵 IGV: {currency} {total_igv:,.2f}\n\n"
            f"📋 **Facturas:** {cantidad}\n"
            f"📈 **Promedio:** {currency} {promedio:,.2f}\n"
        )
        
        await update.message.reply_text(summary)
        
        # Send Excel
        filename = f"resumen_{fecha_inicio_str}_{fecha_fin_str}.xlsx"
        await update.message.reply_document(
            document=excel_file,
            filename=filename,
            caption="📎 Resumen detallado en Excel"
        )
        
        logger.log_info(
            "Resumen generated",
            chat_id=chat_id,
            fecha_inicio=str(fecha_inicio),
            fecha_fin=str(fecha_fin),
            total=total_general,
            cantidad=cantidad
        )
        
    except Exception as e:
        logger.log_error("Error in resumen_command", error=str(e), chat_id=chat_id)
        await update.message.reply_text(f"❌ Error al generar resumen: {str(e)}")


async def proveedores_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle /proveedores [mes] or /proveedores [inicio] [fin]"""
    chat_id = str(update.effective_chat.id)
    
    try:
        if not supabase_service.is_enabled():
            await update.message.reply_text("❌ Servicio de base de datos no disponible.")
            return
        
        # Parse period
        if len(context.args) == 0:
            # Current month
            now = datetime.now()
            fecha_inicio = datetime(now.year, now.month, 1).date()
            if now.month == 12:
                fecha_fin = datetime(now.year + 1, 1, 1).date()
            else:
                fecha_fin = datetime(now.year, now.month + 1, 1).date()
            periodo_str = f"{now.month:02d}-{now.year}"
        elif len(context.args) == 1:
            # MM-YYYY format
            try:
                month, year = context.args[0].split("-")
                month, year = int(month), int(year)
                fecha_inicio = datetime(year, month, 1).date()
                if month == 12:
                    fecha_fin = datetime(year + 1, 1, 1).date()
                else:
                    fecha_fin = datetime(year, month + 1, 1).date()
                periodo_str = f"{month:02d}-{year}"
            except:
                await update.message.reply_text(
                    "❌ Formato incorrecto.\n\n"
                    "Uso: `/proveedores MM-YYYY`\n"
                    "Ejemplo: `/proveedores 01-2026`"
                )
                return
        else:
            # Two dates
            try:
                fecha_inicio = datetime.strptime(context.args[0], "%d-%m-%Y").date()
                fecha_fin = datetime.strptime(context.args[1], "%d-%m-%Y").date()
                periodo_str = f"{context.args[0]} - {context.args[1]}"
            except:
                await update.message.reply_text(
                    "❌ Formato de fecha inválido.\n\n"
                    "Usa: DD-MM-YYYY"
                )
                return
        
        await update.message.reply_text("⏳ Analizando proveedores...")
        
        # Query invoices
        result = supabase_service.client.table("invoices").select("*").eq(
            "company_id", chat_id
        ).gte("invoice_date", str(fecha_inicio)).lt("invoice_date", str(fecha_fin)).is_(
            "deleted_at", "null"
        ).execute()
        
        invoices = result.data if result.data else []
        
        if not invoices:
            await update.message.reply_text(f"📭 No hay facturas en {periodo_str}")
            return
        
        # Group by supplier
        from collections import defaultdict
        suppliers = defaultdict(lambda: {"total": 0, "count": 0, "ruc": ""})
        
        for inv in invoices:
            supplier = inv.get("supplier_name", "Sin nombre")
            total = float(inv.get("total", 0) or 0)
            ruc = inv.get("supplier_ruc", "")
            
            suppliers[supplier]["total"] += total
            suppliers[supplier]["count"] += 1
            if ruc and not suppliers[supplier]["ruc"]:
                suppliers[supplier]["ruc"] = ruc
        
        # Sort by total descending
        sorted_suppliers = sorted(suppliers.items(), key=lambda x: x[1]["total"], reverse=True)
        top_10 = sorted_suppliers[:10]
        
        currency = invoices[0].get("currency", "PEN") if invoices else "PEN"
        total_general = sum(s[1]["total"] for s in sorted_suppliers)
        
        # Generate message
        message = f"🏪 **Top Proveedores {periodo_str}**\n\n"
        
        for idx, (name, data) in enumerate(top_10, 1):
            percentage = (data["total"] / total_general * 100) if total_general > 0 else 0
            message += f"{idx}. **{name}**\n"
            message += f"   💰 {currency} {data['total']:,.2f} ({percentage:.1f}%)\n"
            message += f"   📄 {data['count']} factura(s)\n"
            if data["ruc"]:
                message += f"   🆔 RUC: {data['ruc']}\n"
            message += "\n"
        
        message += f"💰 **Total periodo:** {currency} {total_general:,.2f}"
        
        await update.message.reply_text(message)
        
        logger.log_info(
            "Proveedores analyzed",
            chat_id=chat_id,
            periodo=periodo_str,
            total_suppliers=len(suppliers)
        )
        
    except Exception as e:
        logger.log_error("Error in proveedores_command", error=str(e), chat_id=chat_id)
        await update.message.reply_text(f"❌ Error: {str(e)}")


async def items_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle /items [termino] or /items [mes]"""
    chat_id = str(update.effective_chat.id)
    
    try:
        if not supabase_service.is_enabled():
            await update.message.reply_text("❌ Servicio de base de datos no disponible.")
            return
        
        if len(context.args) == 0:
            await update.message.reply_text(
                "❌ Debes especificar un término de búsqueda o periodo.\n\n"
                "Ejemplos:\n"
                "`/items laptop`\n"
                "`/items 01-2026`"
            )
            return
        
        search_term = " ".join(context.args)
        
        await update.message.reply_text(f"🔍 Buscando items con '{search_term}'...")
        
        # Query invoice_items table
        result = supabase_service.client.table("invoice_items").select(
            "*, invoices!inner(invoice_date, invoice_number, supplier_name, company_id)"
        ).eq("invoices.company_id", chat_id).ilike(
            "description", f"%{search_term}%"
        ).order("invoices.invoice_date", desc=True).limit(50).execute()
        
        items = result.data if result.data else []
        
        if not items:
            await update.message.reply_text(f"📭 No se encontraron items con '{search_term}'")
            return
        
        # Format response
        message = f"💼 **Items encontrados:** {len(items)}\n"
        message += f"🔍 Búsqueda: '{search_term}'\n\n"
        
        total_spent = 0
        for idx, item in enumerate(items[:20], 1):  # Show first 20
            invoice = item.get("invoices", {})
            date = invoice.get("invoice_date", "")
            inv_num = invoice.get("invoice_number", "")
            supplier = invoice.get("supplier_name", "")
            desc = item.get("description", "")
            qty = item.get("quantity", 0) or 0
            unit_price = item.get("unit_price", 0) or 0
            total_price = item.get("total_price", 0) or 0
            
            total_spent += float(total_price)
            
            message += f"{idx}. **{desc[:40]}**\n"
            message += f"   📅 {date} | {inv_num}\n"
            message += f"   🏪 {supplier}\n"
            message += f"   📦 Cant: {qty} x S/ {unit_price:.2f} = S/ {total_price:.2f}\n\n"
            
            if idx >= 10:  # Limit to avoid too long messages
                break
        
        if len(items) > 20:
            message += f"_(Mostrando 20 de {len(items)} resultados)_\n\n"
        
        message += f"💰 **Total gastado:** S/ {total_spent:,.2f}"
        
        await update.message.reply_text(message)
        
        logger.log_info(
            "Items searched",
            chat_id=chat_id,
            search_term=search_term,
            results=len(items)
        )
        
    except Exception as e:
        logger.log_error("Error in items_command", error=str(e), chat_id=chat_id)
        await update.message.reply_text(f"❌ Error: {str(e)}")


async def buscar_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle /buscar [termino]"""
    chat_id = str(update.effective_chat.id)
    
    try:
        if not supabase_service.is_enabled():
            await update.message.reply_text("❌ Servicio de base de datos no disponible.")
            return
        
        if len(context.args) == 0:
            await update.message.reply_text(
                "❌ Debes especificar un término de búsqueda.\n\n"
                "Ejemplos:\n"
                "`/buscar F001-12345`\n"
                "`/buscar Sodimac`\n"
                "`/buscar 20123456789`"
            )
            return
        
        search_term = " ".join(context.args)
        
        await update.message.reply_text(f"🔍 Buscando '{search_term}'...")
        
        # Search in multiple fields
        result = supabase_service.client.table("invoices").select("*").eq(
            "company_id", chat_id
        ).is_("deleted_at", "null").or_(
            f"invoice_number.ilike.%{search_term}%,"
            f"supplier_name.ilike.%{search_term}%,"
            f"supplier_ruc.ilike.%{search_term}%,"
            f"customer_name.ilike.%{search_term}%"
        ).order("invoice_date", desc=True).limit(20).execute()
        
        invoices = result.data if result.data else []
        
        if not invoices:
            await update.message.reply_text(f"📭 No se encontraron facturas con '{search_term}'")
            return
        
        # Format response
        message = f"📄 **Facturas encontradas:** {len(invoices)}\n"
        message += f"🔍 Búsqueda: '{search_term}'\n\n"
        
        for idx, inv in enumerate(invoices[:10], 1):
            date = inv.get("invoice_date", "")
            num = inv.get("invoice_number", "")
            supplier = inv.get("supplier_name", "")
            total = float(inv.get("total", 0) or 0)
            currency = inv.get("currency", "PEN")
            
            message += f"{idx}. **{num or 'Sin número'}**\n"
            message += f"   📅 {date}\n"
            message += f"   🏪 {supplier}\n"
            message += f"   💰 {currency} {total:,.2f}\n\n"
        
        if len(invoices) > 10:
            message += f"_(Mostrando 10 de {len(invoices)} resultados)_"
        
        await update.message.reply_text(message)
        
        logger.log_info(
            "Invoices searched",
            chat_id=chat_id,
            search_term=search_term,
            results=len(invoices)
        )
        
    except Exception as e:
        logger.log_error("Error in buscar_command", error=str(e), chat_id=chat_id)
        await update.message.reply_text(f"❌ Error: {str(e)}")


async def historial_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle /historial [cantidad]"""
    chat_id = str(update.effective_chat.id)
    
    try:
        if not supabase_service.is_enabled():
            await update.message.reply_text("❌ Servicio de base de datos no disponible.")
            return
        
        # Parse limit
        limit = 10
        if len(context.args) > 0:
            try:
                limit = int(context.args[0])
                if limit < 1 or limit > 50:
                    limit = 10
            except:
                limit = 10
        
        await update.message.reply_text(f"📚 Consultando últimas {limit} facturas...")
        
        # Query recent invoices
        result = supabase_service.client.table("invoices").select("*").eq(
            "company_id", chat_id
        ).is_("deleted_at", "null").order("created_at", desc=True).limit(limit).execute()
        
        invoices = result.data if result.data else []
        
        if not invoices:
            await update.message.reply_text("📭 No hay facturas registradas")
            return
        
        # Format response
        message = f"📚 **Últimas {len(invoices)} facturas**\n\n"
        
        for idx, inv in enumerate(invoices, 1):
            date = inv.get("invoice_date", "Sin fecha")
            num = inv.get("invoice_number", "Sin número")
            supplier = inv.get("supplier_name", "Sin proveedor")
            total = float(inv.get("total", 0) or 0)
            currency = inv.get("currency", "PEN")
            
            message += f"{idx}. **{num}**\n"
            message += f"   📅 {date} | 🏪 {supplier}\n"
            message += f"   💰 {currency} {total:,.2f}\n\n"
        
        await update.message.reply_text(message)
        
        logger.log_info(
            "Historial displayed",
            chat_id=chat_id,
            limit=limit,
            results=len(invoices)
        )
        
    except Exception as e:
        logger.log_error("Error in historial_command", error=str(e), chat_id=chat_id)
        await update.message.reply_text(f"❌ Error: {str(e)}")


async def eliminar_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle /eliminar - show recent invoices with delete buttons"""
    chat_id = str(update.effective_chat.id)
    
    try:
        if not supabase_service.is_enabled():
            await update.message.reply_text("❌ Servicio de base de datos no disponible.")
            return
        
        # Query recent invoices
        result = supabase_service.client.table("invoices").select("id, invoice_number, invoice_date, supplier_name, total, currency, created_at").eq(
            "company_id", chat_id
        ).is_("deleted_at", "null").order("created_at", desc=True).limit(5).execute()
        
        invoices = result.data if result.data else []
        
        if not invoices:
            await update.message.reply_text("📭 No hay facturas para eliminar")
            return
        
        # Create buttons for each invoice
        keyboard = []
        for inv in invoices:
            inv_id = inv.get("id")
            num = inv.get("invoice_number", "Sin número")
            date = inv.get("invoice_date", "")
            supplier = inv.get("supplier_name", "")[:10]
            total = float(inv.get("total", 0) or 0)
            created_at = inv.get("created_at", "")
            
            # Format created_at to show date and time
            if created_at:
                try:
                    dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                    fecha_subida = dt.strftime("%d/%m %H:%M")
                except:
                    fecha_subida = ""
            else:
                fecha_subida = ""
            
            button_text = f"🗑️ {num} - {supplier} - S/ {total:.0f} - 📅 {fecha_subida}"
            keyboard.append([InlineKeyboardButton(button_text, callback_data=f"delete_{inv_id}")])
        
        keyboard.append([InlineKeyboardButton("❌ Cancelar", callback_data="delete_cancel")])
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await update.message.reply_text(
            "🗑️ **Selecciona factura a eliminar:**\n\n"
            "⚠️ La factura será eliminada permanentemente de la base de datos.",
            reply_markup=reply_markup
        )
        
        logger.log_info("Delete menu shown", chat_id=chat_id, invoices_count=len(invoices))
        
    except Exception as e:
        logger.log_error("Error in eliminar_command", error=str(e), chat_id=chat_id)
        await update.message.reply_text(f"❌ Error: {str(e)}")


async def estadisticas_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle /estadisticas [mes]"""
    chat_id = str(update.effective_chat.id)
    
    try:
        if not supabase_service.is_enabled():
            await update.message.reply_text("❌ Servicio de base de datos no disponible.")
            return
        
        from collections import defaultdict
        
        # Parse period
        if len(context.args) == 0:
            now = datetime.now()
            month, year = now.month, now.year
        else:
            try:
                month, year = context.args[0].split("-")
                month, year = int(month), int(year)
            except:
                await update.message.reply_text(
                    "❌ Formato incorrecto. Usa: `/estadisticas MM-YYYY`\n"
                    "Ejemplo: `/estadisticas 01-2026`"
                )
                return
        
        fecha_inicio = datetime(year, month, 1).date()
        if month == 12:
            fecha_fin = datetime(year + 1, 1, 1).date()
        else:
            fecha_fin = datetime(year, month + 1, 1).date()
        
        await update.message.reply_text("📊 Generando estadísticas...")
        
        # Query invoices
        result = supabase_service.client.table("invoices").select("*").eq(
            "company_id", chat_id
        ).gte("invoice_date", str(fecha_inicio)).lt("invoice_date", str(fecha_fin)).is_(
            "deleted_at", "null"
        ).execute()
        
        invoices = result.data if result.data else []
        
        if not invoices:
            await update.message.reply_text(f"📭 No hay facturas en {month:02d}-{year}")
            return
        
        # Calculate statistics
        total = sum(float(inv.get("total", 0) or 0) for inv in invoices)
        cantidad = len(invoices)
        promedio = total / cantidad if cantidad > 0 else 0
        currency = invoices[0].get("currency", "PEN") if invoices else "PEN"
        
        # Group by day
        by_day = defaultdict(float)
        for inv in invoices:
            date = inv.get("invoice_date", "")
            if date:
                by_day[date] += float(inv.get("total", 0) or 0)
        
        max_day = max(by_day.items(), key=lambda x: x[1]) if by_day else (None, 0)
        
        # Top suppliers
        suppliers = defaultdict(float)
        for inv in invoices:
            supplier = inv.get("supplier_name", "Sin nombre")
            suppliers[supplier] += float(inv.get("total", 0) or 0)
        
        top_3 = sorted(suppliers.items(), key=lambda x: x[1], reverse=True)[:3]
        
        # Format message
        message = f"📊 **Estadísticas {month:02d}/{year}**\n\n"
        message += f"💰 **Total gastado:** {currency} {total:,.2f}\n"
        message += f"📄 **Facturas:** {cantidad}\n"
        message += f"📈 **Promedio:** {currency} {promedio:,.2f}\n\n"
        
        if max_day[0]:
            message += f"🔝 **Día con más gastos:**\n"
            message += f"   {max_day[0]} - {currency} {max_day[1]:,.2f}\n\n"
        
        if top_3:
            message += f"🏪 **Top 3 Proveedores:**\n"
            for idx, (name, amount) in enumerate(top_3, 1):
                message += f"   {idx}. {name} - {currency} {amount:,.2f}\n"
        
        await update.message.reply_text(message)
        
        logger.log_info(
            "Estadisticas generated",
            chat_id=chat_id,
            month=month,
            year=year,
            total=total
        )
        
    except Exception as e:
        logger.log_error("Error in estadisticas_command", error=str(e), chat_id=chat_id)
        await update.message.reply_text(f"❌ Error: {str(e)}")


async def comparar_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle /comparar [mes1] [mes2]"""
    chat_id = str(update.effective_chat.id)
    
    try:
        if not supabase_service.is_enabled():
            await update.message.reply_text("❌ Servicio de base de datos no disponible.")
            return
        
        if len(context.args) < 2:
            await update.message.reply_text(
                "❌ Debes especificar dos periodos.\n\n"
                "Uso: `/comparar MM-YYYY MM-YYYY`\n"
                "Ejemplo: `/comparar 01-2026 12-2025`"
            )
            return
        
        # Parse periods
        try:
            m1, y1 = context.args[0].split("-")
            m2, y2 = context.args[1].split("-")
            m1, y1, m2, y2 = int(m1), int(y1), int(m2), int(y2)
        except:
            await update.message.reply_text("❌ Formato inválido. Usa: MM-YYYY")
            return
        
        # Period 1
        fecha1_inicio = datetime(y1, m1, 1).date()
        if m1 == 12:
            fecha1_fin = datetime(y1 + 1, 1, 1).date()
        else:
            fecha1_fin = datetime(y1, m1 + 1, 1).date()
        
        # Period 2
        fecha2_inicio = datetime(y2, m2, 1).date()
        if m2 == 12:
            fecha2_fin = datetime(y2 + 1, 1, 1).date()
        else:
            fecha2_fin = datetime(y2, m2 + 1, 1).date()
        
        await update.message.reply_text("📊 Comparando periodos...")
        
        # Query period 1
        result1 = supabase_service.client.table("invoices").select("total, currency").eq(
            "company_id", chat_id
        ).gte("invoice_date", str(fecha1_inicio)).lt("invoice_date", str(fecha1_fin)).is_(
            "deleted_at", "null"
        ).execute()
        
        # Query period 2
        result2 = supabase_service.client.table("invoices").select("total, currency").eq(
            "company_id", chat_id
        ).gte("invoice_date", str(fecha2_inicio)).lt("invoice_date", str(fecha2_fin)).is_(
            "deleted_at", "null"
        ).execute()
        
        inv1 = result1.data if result1.data else []
        inv2 = result2.data if result2.data else []
        
        total1 = sum(float(i.get("total", 0) or 0) for i in inv1)
        total2 = sum(float(i.get("total", 0) or 0) for i in inv2)
        count1 = len(inv1)
        count2 = len(inv2)
        
        currency = inv1[0].get("currency", "PEN") if inv1 else "PEN"
        
        # Calculate difference
        diff = total1 - total2
        diff_pct = ((total1 - total2) / total2 * 100) if total2 > 0 else 0
        
        # Format message
        message = f"📊 **Comparación de Periodos**\n\n"
        message += f"📅 **Periodo 1:** {m1:02d}/{y1}\n"
        message += f"   💰 {currency} {total1:,.2f} ({count1} facturas)\n\n"
        message += f"📅 **Periodo 2:** {m2:02d}/{y2}\n"
        message += f"   💰 {currency} {total2:,.2f} ({count2} facturas)\n\n"
        message += f"📈 **Diferencia:**\n"
        
        if diff > 0:
            message += f"   ⬆️ +{currency} {diff:,.2f} (+{diff_pct:.1f}%)\n"
            message += f"   ⚠️ Gastaste MÁS en {m1:02d}/{y1}"
        elif diff < 0:
            message += f"   ⬇️ {currency} {diff:,.2f} ({diff_pct:.1f}%)\n"
            message += f"   ✅ Gastaste MENOS en {m1:02d}/{y1}"
        else:
            message += f"   ➡️ Sin cambios"
        
        await update.message.reply_text(message)
        
        logger.log_info(
            "Periods compared",
            chat_id=chat_id,
            period1=f"{m1:02d}-{y1}",
            period2=f"{m2:02d}-{y2}",
            diff=diff
        )
        
    except Exception as e:
        logger.log_error("Error in comparar_command", error=str(e), chat_id=chat_id)
        await update.message.reply_text(f"❌ Error: {str(e)}")


async def cancelar_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle /cancelar - cancel field editing"""
    if 'editing_field' in context.user_data:
        del context.user_data['editing_field']
        del context.user_data['editing_validation_key']
        await update.message.reply_text("❌ Edición cancelada")
    else:
        await update.message.reply_text("⚠️ No hay ninguna edición en curso")


async def handle_text_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle text messages for field editing"""
    # Check if user is editing a field
    if 'editing_field' not in context.user_data:
        return
    
    field_name = context.user_data.get('editing_field')
    validation_key = context.user_data.get('editing_validation_key')
    new_value = update.message.text.strip()
    
    validation_data = pending_validations.get(validation_key)
    if not validation_data:
        await update.message.reply_text("⚠️ Sesión expirada. Por favor, procesa la factura nuevamente.")
        del context.user_data['editing_field']
        del context.user_data['editing_validation_key']
        return
    
    invoice = validation_data["invoice"]
    
    # Update the field
    try:
        if field_name == "invoice_number":
            invoice.invoice_number = new_value
        elif field_name == "invoice_date":
            invoice.invoice_date = new_value
        elif field_name == "supplier_name":
            invoice.supplier_name = new_value
        elif field_name == "supplier_ruc":
            invoice.supplier_ruc = new_value
        elif field_name == "customer_name":
            invoice.customer_name = new_value
        elif field_name == "subtotal":
            invoice.subtotal = float(new_value)
        elif field_name == "tax":
            invoice.tax = float(new_value)
        elif field_name == "total":
            invoice.total = float(new_value)
        
        await update.message.reply_text(f"✅ Campo actualizado: {new_value}")
        
        # Clear editing state
        del context.user_data['editing_field']
        del context.user_data['editing_validation_key']
        
        # Show updated validation summary
        items_summary = ""
        if invoice.items and len(invoice.items) > 0:
            items_summary = f"\n📦 **Items:** {len(invoice.items)} producto(s)"
            for idx, item in enumerate(invoice.items[:3], 1):
                items_summary += f"\n   {idx}. {item.description or 'Sin descripción'}"
                items_summary += f" - {item.quantity}x S/ {item.unit_price or 0:.2f}"
            if len(invoice.items) > 3:
                items_summary += f"\n   ... y {len(invoice.items) - 3} más"
        
        validation_message = (
            f"📋 **Datos actualizados**\n\n"
            f"📄 **Nro. Factura:** {invoice.invoice_number or 'N/A'}\n"
            f"📅 **Fecha:** {invoice.invoice_date or 'N/A'}\n"
            f"🏪 **Proveedor:** {invoice.supplier_name or 'N/A'}\n"
            f"🆔 **RUC:** {invoice.supplier_ruc or 'N/A'}\n"
            f"👤 **Cliente:** {invoice.customer_name or 'N/A'}\n\n"
            f"💰 **Subtotal:** {invoice.currency or 'PEN'} {invoice.subtotal or 0:.2f}\n"
            f"📊 **IGV:** {invoice.currency or 'PEN'} {invoice.tax or 0:.2f}\n"
            f"💵 **Total:** {invoice.currency or 'PEN'} {invoice.total or 0:.2f}\n"
            f"{items_summary}\n\n"
            f"✅ **¿Los datos son correctos?**"
        )
        
        keyboard = [
            [
                InlineKeyboardButton("✅ Guardar", callback_data=f"save_{validation_key}"),
                InlineKeyboardButton("✏️ Editar", callback_data=f"edit_{validation_key}")
            ],
            [InlineKeyboardButton("❌ Cancelar", callback_data=f"cancel_{validation_key}")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await update.message.reply_text(validation_message, reply_markup=reply_markup)
        
    except ValueError:
        await update.message.reply_text("❌ Valor inválido. Intenta nuevamente o usa /cancelar")
    except Exception as e:
        logger.log_error("Error updating field", error=str(e))
        await update.message.reply_text(f"❌ Error: {str(e)}")


def main():
    """Start the bot"""
    if not settings.telegram_bot_token:
        logger.log_error("TELEGRAM_BOT_TOKEN not configured in .env")
        print("❌ Error: TELEGRAM_BOT_TOKEN no está configurado en el archivo .env")
        print("Por favor, obtén tu token de @BotFather y agrégalo al .env")
        return
    
    logger.log_info("Starting Telegram bot...")
    
    # Create application
    application = Application.builder().token(settings.telegram_bot_token).build()
    
    # Register handlers
    application.add_handler(CommandHandler("start", start_command))
    application.add_handler(CommandHandler("ayuda", help_command))
    application.add_handler(CommandHandler("resumen", resumen_command))
    application.add_handler(CommandHandler("proveedores", proveedores_command))
    application.add_handler(CommandHandler("items", items_command))
    application.add_handler(CommandHandler("buscar", buscar_command))
    application.add_handler(CommandHandler("historial", historial_command))
    application.add_handler(CommandHandler("eliminar", eliminar_command))
    application.add_handler(CommandHandler("estadisticas", estadisticas_command))
    application.add_handler(CommandHandler("comparar", comparar_command))
    application.add_handler(CommandHandler("cancelar", cancelar_command))
    application.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    application.add_handler(MessageHandler(filters.Document.PDF, handle_document))
    application.add_handler(CallbackQueryHandler(button_callback))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text_message))
    
    # Start bot
    logger.log_info("Bot is running... Press Ctrl+C to stop.")
    print("🤖 Bot iniciado correctamente!")
    print("📱 Envía /start al bot para comenzar")
    print("⏹️  Presiona Ctrl+C para detener")
    
    application.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
