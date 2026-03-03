from fastapi import APIRouter, UploadFile, File, HTTPException, status, Form, Query
from fastapi.responses import StreamingResponse, JSONResponse
from app.services.zip_handler import ZipHandler
from app.services.processing_orchestrator import orchestrator
from app.services.storage_service import storage_service
from app.services.supabase_service import supabase_service
from app.services.gemini_service import gemini_service
from app.schemas.invoice_schema import ProcessingResponse, InvoiceSchema, SaveInvoiceRequest, SaveInvoiceResponse
from app.utils.logger import setup_logger
from pydantic import BaseModel, HttpUrl, model_validator
from typing import Optional
import io
import base64
import httpx
import uuid
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta

logger = setup_logger("upload_router")

router = APIRouter()

zip_handler = ZipHandler()


class MediaUrlRequest(BaseModel):
    """Request model for processing media from URL (WhatsApp API)
    Acepta tanto camelCase (mediaUrl) como snake_case (media_url).
    """
    mediaUrl: Optional[str] = None
    phoneNumber: Optional[str] = None
    callbackUrl: Optional[str] = None
    mimeType: Optional[str] = None
    access_token: Optional[str] = None
    filename: Optional[str] = None

    @model_validator(mode="before")
    @classmethod
    def normalize_fields(cls, data: dict) -> dict:
        """Normaliza snake_case → camelCase para compatibilidad con clientes externos."""
        if isinstance(data, dict):
            if not data.get("mediaUrl") and data.get("media_url"):
                data["mediaUrl"] = data["media_url"]
            if not data.get("phoneNumber") and data.get("phone_number"):
                data["phoneNumber"] = data["phone_number"]
            if not data.get("callbackUrl") and data.get("callback_url"):
                data["callbackUrl"] = data["callback_url"]
            # También acepta 'url' directamente
            if not data.get("mediaUrl") and data.get("url"):
                data["mediaUrl"] = data["url"]
        return data

    @model_validator(mode="after")
    def require_media_url(self) -> "MediaUrlRequest":
        if not self.mediaUrl:
            raise ValueError("Se requiere 'mediaUrl' (o 'media_url' / 'url')")
        return self



async def download_from_url(url: str, access_token: Optional[str] = None) -> bytes:
    """Download file from URL with optional authentication for WhatsApp API"""
    headers = {}
    if access_token:
        headers["Authorization"] = f"Bearer {access_token}"
    
    async with httpx.AsyncClient(timeout=30.0) as client:
        response = await client.get(url, headers=headers)
        response.raise_for_status()
        return response.content


async def send_callback(callback_url: str, payload: dict) -> None:
    """Send callback notification to the provided URL"""
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            response = await client.post(
                callback_url,
                json=payload,
                headers={"Content-Type": "application/json"}
            )
            response.raise_for_status()
            logger.log_info(
                "Callback sent successfully",
                callback_url=callback_url,
                status_code=response.status_code
            )
    except Exception as e:
        logger.log_error(
            "Failed to send callback",
            callback_url=callback_url,
            error=str(e)
        )

def create_excel_from_results(processing_response: ProcessingResponse, filename: str) -> io.BytesIO:
    """Convert processing results to Excel file"""
    wb = Workbook()
    
    # Sheet 1: Invoices Summary
    ws_summary = wb.active
    ws_summary.title = "Resumen Facturas"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Headers for summary
    headers = [
        "Nro. Factura", "Fecha", "Proveedor", "RUC Proveedor", 
        "Cliente", "CC/Placa", "Tipo Documento", "Tipo Costo", "Subtotal", 
        "IGV", "Total", "Moneda", "Confianza", "Archivo Origen", "Secuencia"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    for row_idx, invoice in enumerate(processing_response.results, 2):
        ws_summary.cell(row=row_idx, column=1, value=invoice.invoice_number or "")
        ws_summary.cell(row=row_idx, column=2, value=invoice.invoice_date or "")
        ws_summary.cell(row=row_idx, column=3, value=invoice.supplier_name or "")
        ws_summary.cell(row=row_idx, column=4, value=invoice.supplier_ruc or "")
        ws_summary.cell(row=row_idx, column=5, value=invoice.customer_name or "")
        ws_summary.cell(row=row_idx, column=6, value=getattr(invoice, 'cc_or_placa', None) or "")
        ws_summary.cell(row=row_idx, column=7, value=getattr(invoice, 'document_type', None) or "")
        ws_summary.cell(row=row_idx, column=8, value=getattr(invoice, 'tipo_costo', None) or "")
        ws_summary.cell(row=row_idx, column=9, value=invoice.subtotal)
        ws_summary.cell(row=row_idx, column=10, value=invoice.tax)
        ws_summary.cell(row=row_idx, column=11, value=invoice.total)
        ws_summary.cell(row=row_idx, column=12, value=invoice.currency or "")
        ws_summary.cell(row=row_idx, column=13, value=invoice.confidence_score)
        ws_summary.cell(row=row_idx, column=14, value=invoice.source_file)
        ws_summary.cell(row=row_idx, column=15, value=invoice.sequence_id)
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws_summary.column_dimensions[ws_summary.cell(row=1, column=col).column_letter].width = 15
    
    # Sheet 2: Items Detail
    ws_items = wb.create_sheet("Detalle Items")
    
    item_headers = [
        "Nro. Factura", "Fecha Factura", "Proveedor", "RUC Proveedor",
        "Cliente", "Tipo Documento", "CC/Placa", "Moneda", "Archivo Origen", 
        "Item #", "Descripción", "Cantidad", "Unidad", "Precio Unit.", "Total Item"
    ]
    
    for col, header in enumerate(item_headers, 1):
        cell = ws_items.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Items data
    item_row = 2
    for invoice in processing_response.results:
        if invoice.items:
            for item_idx, item in enumerate(invoice.items, 1):
                ws_items.cell(row=item_row, column=1, value=invoice.invoice_number or "")
                ws_items.cell(row=item_row, column=2, value=invoice.invoice_date or "")
                ws_items.cell(row=item_row, column=3, value=invoice.supplier_name or "")
                ws_items.cell(row=item_row, column=4, value=invoice.supplier_ruc or "")
                ws_items.cell(row=item_row, column=5, value=invoice.customer_name or "")
                ws_items.cell(row=item_row, column=6, value=getattr(invoice, 'document_type', None) or "")
                ws_items.cell(row=item_row, column=7, value=getattr(invoice, 'cc_or_placa', None) or "")
                ws_items.cell(row=item_row, column=8, value=invoice.currency or "")
                ws_items.cell(row=item_row, column=9, value=invoice.source_file)
                ws_items.cell(row=item_row, column=10, value=item_idx)
                ws_items.cell(row=item_row, column=11, value=item.description or "")
                ws_items.cell(row=item_row, column=12, value=item.quantity)
                ws_items.cell(row=item_row, column=13, value=item.unit or "")
                ws_items.cell(row=item_row, column=14, value=item.unit_price)
                ws_items.cell(row=item_row, column=15, value=item.total_price)
                item_row += 1
    
    # Auto-adjust column widths
    for col in range(1, len(item_headers) + 1):
        ws_items.column_dimensions[ws_items.cell(row=1, column=col).column_letter].width = 18
    
    # Sheet 3: Errors
    if processing_response.errors:
        ws_errors = wb.create_sheet("Errores")
        
        error_headers = ["Archivo", "Secuencia", "Tipo", "Error"]
        for col, header in enumerate(error_headers, 1):
            cell = ws_errors.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for row_idx, error in enumerate(processing_response.errors, 2):
            ws_errors.cell(row=row_idx, column=1, value=error.get("file", ""))
            ws_errors.cell(row=row_idx, column=2, value=error.get("sequence_id", ""))
            ws_errors.cell(row=row_idx, column=3, value=error.get("file_type", ""))
            ws_errors.cell(row=row_idx, column=4, value=error.get("error", ""))
        
        for col in range(1, len(error_headers) + 1):
            ws_errors.column_dimensions[ws_errors.cell(row=1, column=col).column_letter].width = 20
    
    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file

@router.post("/invoices", response_model=SaveInvoiceResponse)
async def save_invoice(request: SaveInvoiceRequest):
    """
    Save invoice data to Supabase database (called from WhatsApp API)
    
    Saves to 2 tables:
    1. **invoices** - Main invoice data
    2. **invoice_items** - Individual line items
    
    Example request body:
    ```json
    {
      "supplier_ruc": "20123456789",
      "supplier_name": "EMPRESA S.A.C.",
      "customer_ruc": "20987654321",
      "customer_name": "CLIENTE ABC",
      "invoice_date": "2026-02-18",
      "invoice_number": "F001-00012345",
      "subtotal": 2118.64,
      "igv": 381.36,
      "total": 2500.00,
      "currency": "PEN",
      "items": [
        {
          "descripcion": "Producto A",
          "cantidad": 2,
          "precioUnitario": 500.25,
          "total": 1000.50
        }
      ],
      "phoneNumber": "51987654321",
      "job_id": "job_1708272000000"
    }
    ```
    """
    logger.log_info(
        "="*80
    )
    logger.log_info(
        "🔵 POST /invoices - REQUEST RECEIVED",
        phone_number=request.phoneNumber,
        invoice_number=request.invoice_number,
        supplier_name=request.supplier_name,
        supplier_ruc=request.supplier_ruc,
        customer_name=request.customer_name,
        customer_ruc=request.customer_ruc,
        invoice_date=request.invoice_date,
        subtotal=request.subtotal,
        igv=request.igv,
        total=request.total,
        currency=request.currency,
        items_count=len(request.items) if request.items else 0,
        job_id=request.job_id
    )
    
    # Log full request payload for debugging
    import json
    logger.log_info(
        "📦 FULL REQUEST PAYLOAD",
        payload=json.dumps(request.dict(), indent=2, default=str)
    )
    
    try:
        # Check if Supabase is enabled
        logger.log_info("🔍 Step 1: Checking if Supabase is enabled...")
        if not supabase_service.is_enabled():
            logger.log_error("❌ Supabase service is not enabled")
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail="Database service is not available"
            )
        logger.log_info("✅ Supabase is enabled")
        
        # Convert request to dict (exclude metadata fields)
        logger.log_info("🔍 Step 2: Converting request to dict...")
        invoice_dict = request.dict(exclude={'phoneNumber', 'createdAt', 'job_id'}, exclude_none=False)
        
        logger.log_info(
            "✅ Request converted to dict",
            invoice_dict_keys=list(invoice_dict.keys())
        )
        logger.log_info(
            "📋 INVOICE DICT DATA",
            invoice_dict=json.dumps(invoice_dict, indent=2, default=str)
        )
        
        logger.log_info(
            "🔍 Step 3: About to call save_single_invoice...",
            phone_number=request.phoneNumber,
            job_id=request.job_id
        )
        
        # Save to both Supabase tables: invoices and invoice_items
        invoice_id = await supabase_service.save_single_invoice(
            invoice_data=invoice_dict,
            phone_number=request.phoneNumber,
            job_id=request.job_id
        )
        
        logger.log_info(
            "🔍 Step 4: save_single_invoice returned",
            invoice_id=invoice_id,
            invoice_id_type=type(invoice_id).__name__
        )
        
        if not invoice_id:
            logger.log_error(
                "❌ save_single_invoice returned None or falsy value",
                phone_number=request.phoneNumber,
                invoice_number=request.invoice_number,
                returned_value=invoice_id
            )
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to save invoice to database"
            )
        
        logger.log_info(
            "✅ SUCCESS: Invoice saved to Supabase",
            invoice_id=invoice_id,
            phone_number=request.phoneNumber,
            invoice_number=request.invoice_number,
            tables_updated=["invoices", "invoice_items"],
            items_saved=len(request.items) if request.items else 0
        )
        logger.log_info(
            "="*80
        )
        
        return SaveInvoiceResponse(
            id=invoice_id,
            success=True,
            message="Invoice saved successfully"
        )
        
    except HTTPException as http_exc:
        logger.log_error(
            "❌ HTTP EXCEPTION in POST /invoices",
            status_code=http_exc.status_code,
            detail=http_exc.detail,
            phone_number=request.phoneNumber,
            invoice_number=request.invoice_number
        )
        logger.log_info("="*80)
        raise
        
    except Exception as e:
        logger.log_error(
            "❌ UNEXPECTED ERROR in POST /invoices",
            error=str(e),
            error_type=type(e).__name__,
            phone_number=request.phoneNumber,
            invoice_number=request.invoice_number
        )
        import traceback
        logger.log_error(
            "📍 FULL TRACEBACK",
            traceback=traceback.format_exc()
        )
        logger.log_info("="*80)
        
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Internal server error: {str(e)}"
        )

@router.post("/process-zip")
async def process_zip_file(file: UploadFile = File(...)):
    """
    Process a ZIP file from WhatsApp and extract invoice data
    
    - **file**: ZIP file containing WhatsApp exported media
    
    Returns structured invoice data extracted from images, PDFs, and text files.
    Files are processed in chronological order based on WhatsApp timestamps.
    """

    logger.log_info(
        "ZIP file upload started",
        filename=file.filename,
        content_type=file.content_type
    )

    try:
        # Validate file type
        if not file.filename or not file.filename.lower().endswith('.zip'):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="File must be a ZIP file"
            )
        
        # Read file content
        file_content = await file.read()
        
        logger.log_info(
            "ZIP file read successfully",
            filename=file.filename,
            size_bytes=len(file_content)
        )
        
        # Upload ZIP to Cloud Storage (backup y audit trail)
        storage_metadata = storage_service.upload_zip(
            file_content, 
            file.filename,
            company_id=None  # TODO: extraer de headers o auth cuando implementes multi-tenancy
        )
        
        if storage_metadata:
            logger.log_info(
                "ZIP backed up to Cloud Storage",
                blob_path=storage_metadata["blob_path"],
                signed_url=storage_metadata["signed_url"]
            )
        else:
            # No falla si no se puede subir a Storage (degrada gracefully)
            logger.log_warning(
                "Failed to backup ZIP to Cloud Storage, continuing with processing",
                filename=file.filename
            )
        
        # Phase 1: Extract and validate ZIP
        success, extraction_result = zip_handler.extract_zip_files(file_content, file.filename)
        
        if not success:
            logger.log_error(
                "ZIP extraction failed",
                filename=file.filename,
                error=extraction_result.get("error")
            )
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"ZIP processing failed: {extraction_result.get('error')}"
            )
        
        extracted_files = extraction_result["extracted_files"]
        ignored_files = extraction_result["ignored_files"]
        temp_dir_obj = extraction_result.get("temp_dir_obj")
        
        if not extracted_files:
            # Cleanup temp directory if no files to process
            if temp_dir_obj:
                temp_dir_obj.cleanup()
            
            logger.log_warning(
                "No processable files found in ZIP",
                filename=file.filename,
                total_ignored=len(ignored_files)
            )
            
            return ProcessingResponse(
                results=[],
                errors=[{"error": "No processable files found in ZIP"}],
                total_processed=0,
                success_count=0
            )
        
        logger.log_info(
            "ZIP extraction completed",
            filename=file.filename,
            extracted_files=len(extracted_files),
            ignored_files=len(ignored_files)
        )
        
        # Phases 2-6: Process files through the pipeline
        try:
            processing_response = await orchestrator.process_extracted_files(extracted_files)
        finally:
            # Cleanup temp directory after processing
            if temp_dir_obj:
                temp_dir_obj.cleanup()
                logger.log_info("Temporary directory cleaned up", filename=file.filename)
        
        logger.log_info(
            "ZIP processing pipeline completed",
            filename=file.filename,
            total_processed=processing_response.total_processed,
            success_count=processing_response.success_count,
            error_count=len(processing_response.errors)
        )
        
        # Generate Excel file
        excel_file = create_excel_from_results(processing_response, file.filename)
        
        # Upload Excel to Cloud Storage
        excel_metadata = storage_service.upload_excel(
            excel_file,
            file.filename,
            company_id=None  # TODO: extraer de headers o auth
        )
        
        # Save to Supabase database
        record_id = None
        if supabase_service.is_enabled():
            try:
                # 1. Save processing record
                record_id = await supabase_service.save_processing_record(
                    chat_id="anonymous",  # TODO: extraer de headers/auth cuando implementes multi-tenancy
                    zip_blob_path=storage_metadata["blob_path"] if storage_metadata else None,
                    excel_blob_path=excel_metadata["blob_path"] if excel_metadata else None,
                    total_files=processing_response.total_processed,
                    success_files=processing_response.success_count,
                    error_files=len(processing_response.errors)
                )
                
                # 2. Save invoices to database
                if record_id and processing_response.results:
                    await supabase_service.save_invoices_batch(
                        invoices=processing_response.results,
                        chat_id="anonymous",  # TODO: extraer de headers/auth
                        record_id=record_id
                    )
                
                logger.log_info(
                    "Data saved to Supabase",
                    record_id=record_id,
                    invoices_count=len(processing_response.results)
                )
                
            except Exception as db_error:
                # Database errors don't fail the request (graceful degradation)
                logger.log_error(
                    "Failed to save to Supabase (continuing with Excel response)",
                    error=str(db_error),
                    filename=file.filename
                )
        else:
            logger.log_info("Supabase integration disabled, skipping database save")
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"facturas_{timestamp}.xlsx"
        
        # Prepare response headers
        response_headers = {
            "Content-Disposition": f"attachment; filename={excel_filename}",
            "X-Total-Files": str(len(extracted_files)),
            "X-Success-Files": str(processing_response.success_count),
            "X-Error-Files": str(len(processing_response.errors))
        }
        
        # Add Cloud Storage URLs to headers if upload succeeded
        if excel_metadata:
            response_headers["X-Storage-Excel-URL"] = excel_metadata["signed_url"]
            logger.log_info(
                "Excel saved to Cloud Storage",
                blob_path=excel_metadata["blob_path"],
                excel_url=excel_metadata["signed_url"]
            )
        
        if storage_metadata:
            response_headers["X-Storage-Zip-URL"] = storage_metadata["signed_url"]
        
        # Reset BytesIO position for streaming
        excel_file.seek(0)
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=response_headers
        )
        
    except HTTPException:
        # Re-raise HTTP exceptions
        raise
        
    except Exception as e:
        logger.log_error(
            "Unexpected error during ZIP processing",
            filename=file.filename if file else "unknown",
            error=str(e)
        )
        
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Internal server error: {str(e)}"
        )


@router.post("/process-image")
async def process_image(file: Optional[UploadFile] = File(None)):
    """
    Process a single image file and extract invoice data
    
    - **file**: Image file (JPG, JPEG, PNG)
    
    Returns structured invoice data extracted from the image.
    """
    if not file:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No file provided. Use 'file' field in form-data or use /process-image-url for URL"
        )
    
    logger.log_info(
        "Image file upload started",
        filename=file.filename,
        content_type=file.content_type
    )
    
    try:
        # Validate file type
        if not file.filename:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Filename is required"
            )
        
        file_ext = file.filename.lower().split('.')[-1]
        if file_ext not in ['jpg', 'jpeg', 'png']:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"File must be an image (JPG, JPEG, PNG). Got: {file_ext}"
            )
        
        # Read file content
        file_content = await file.read()
        
        logger.log_info(
            "Image file read successfully",
            filename=file.filename,
            size_bytes=len(file_content)
        )
        
        # Convert to base64
        image_base64 = base64.b64encode(file_content).decode('utf-8')
        
        # Process with Gemini
        result = await gemini_service.extract_invoice_from_image(
            image_base64=image_base64,
            filename=file.filename,
            sequence_id=1
        )
        
        logger.log_info(
            "Image processing completed",
            filename=file.filename,
            success=result.get('success', False)
        )
        
        return result
        
    except HTTPException:
        raise
        
    except Exception as e:
        logger.log_error(
            "Unexpected error during image processing",
            filename=file.filename if file else "unknown",
            error=str(e)
        )
        
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Internal server error: {str(e)}"
        )


@router.post("/process-image-url")
async def process_image_url(request: MediaUrlRequest):
    """
    Process an image from URL (WhatsApp API) and extract invoice data
    
    - **mediaUrl**: URL of the image to process
    - **phoneNumber**: Phone number of the sender (optional, for logging)
    - **callbackUrl**: URL to send callback when processing is complete (optional)
    - **mimeType**: MIME type of the media (optional)
    - **access_token**: Optional access token for authenticated requests (WhatsApp API)
    - **filename**: Optional filename for logging
    
    Returns structured invoice data extracted from the image.
    """
    filename = request.filename or f"invoice_{request.phoneNumber or 'unknown'}.jpg"
    
    logger.log_info(
        "Image URL processing started",
        url=request.mediaUrl,
        filename=filename,
        phone_number=request.phoneNumber
    )
    
    try:
        # Download file from URL
        file_content = await download_from_url(request.mediaUrl, request.access_token)
        
        logger.log_info(
            "Image downloaded from URL",
            filename=filename,
            size_bytes=len(file_content),
            phone_number=request.phoneNumber
        )
        
        # Convert to base64
        image_base64 = base64.b64encode(file_content).decode('utf-8')
        
        # Process with Gemini
        result = await gemini_service.extract_invoice_from_image(
            image_base64=image_base64,
            filename=filename,
            sequence_id=1
        )
        
        print("resulttttttttttttt", result)
        logger.log_info(
            "Image URL processing completed",
            filename=filename,
            phone_number=request.phoneNumber,
            success=result.get('success', False)
        )
        
        # Prepare response payload
        job_id = f"job_{int(datetime.now().timestamp() * 1000)}"
        is_success = result.get('success', False)
        
        response_payload = {
            "jobId": job_id,
            "status": "success" if is_success else "error",
            "phoneNumber": request.phoneNumber,
            "data": result.get('invoice_data') if is_success else None,
            "error": result.get('error') if not is_success else None
        }
        
        # Send callback if URL provided
        if request.callbackUrl:
            await send_callback(request.callbackUrl, response_payload)
        
        # Return response
        return response_payload
        
    except httpx.HTTPError as e:
        logger.log_error(
            "Failed to download image from URL",
            url=request.mediaUrl,
            error=str(e)
        )
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to download image: {str(e)}"
        )
        
    except Exception as e:
        logger.log_error(
            "Unexpected error during image URL processing",
            url=request.mediaUrl,
            error=str(e)
        )
        
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Internal server error: {str(e)}"
        )


@router.post("/process-pdf")
async def process_pdf(file: Optional[UploadFile] = File(None)):
    """
    Process a single PDF file and extract invoice data
    
    - **file**: PDF file
    
    Returns structured invoice data extracted from the PDF.
    """
    if not file:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No file provided. Use 'file' field in form-data or use /process-pdf-url for URL"
        )
    
    logger.log_info(
        "PDF file upload started",
        filename=file.filename,
        content_type=file.content_type
    )
    
    try:
        # Validate file type
        if not file.filename:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Filename is required"
            )
        
        if not file.filename.lower().endswith('.pdf'):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="File must be a PDF"
            )
        
        # Read file content
        file_content = await file.read()
        
        logger.log_info(
            "PDF file read successfully",
            filename=file.filename,
            size_bytes=len(file_content)
        )
        
        # Convert to base64
        pdf_base64 = base64.b64encode(file_content).decode('utf-8')
        
        # Process with Gemini
        result = await gemini_service.extract_invoice_from_pdf(
            pdf_base64=pdf_base64,
            filename=file.filename,
            sequence_id=1
        )
        
        logger.log_info(
            "PDF processing completed",
            filename=file.filename,
            success=result.get('success', False)
        )
        
        return result
        
    except HTTPException:
        raise
        
    except Exception as e:
        logger.log_error(
            "Unexpected error during PDF processing",
            filename=file.filename if file else "unknown",
            error=str(e)
        )
        
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Internal server error: {str(e)}"
        )


@router.post("/process-pdf-url")
async def process_pdf_url(request: MediaUrlRequest):
    """
    Process a PDF from URL (WhatsApp API) and extract invoice data
    
    - **mediaUrl**: URL of the PDF to process
    - **phoneNumber**: Phone number of the sender (optional, for logging)
    - **callbackUrl**: URL to send callback when processing is complete (optional)
    - **mimeType**: MIME type of the media (optional)
    - **access_token**: Optional access token for authenticated requests (WhatsApp API)
    - **filename**: Optional filename for logging
    
    Returns structured invoice data extracted from the PDF.
    """
    filename = request.filename or f"invoice_{request.phoneNumber or 'unknown'}.pdf"
    
    logger.log_info(
        "PDF URL processing started",
        url=request.mediaUrl,
        filename=filename,
        phone_number=request.phoneNumber
    )
    
    try:
        # Download file from URL
        file_content = await download_from_url(request.mediaUrl, request.access_token)
        
        logger.log_info(
            "PDF downloaded from URL",
            filename=filename,
            size_bytes=len(file_content),
            phone_number=request.phoneNumber
        )
        
        # Convert to base64
        pdf_base64 = base64.b64encode(file_content).decode('utf-8')
        
        # Process with Gemini
        result = await gemini_service.extract_invoice_from_pdf(
            pdf_base64=pdf_base64,
            filename=filename,
            sequence_id=1
        )
        
        logger.log_info(
            "PDF URL processing completed",
            filename=filename,
            phone_number=request.phoneNumber,
            success=result.get('success', False)
        )
        
        # Prepare response payload
        job_id = f"job_{int(datetime.now().timestamp() * 1000)}"
        is_success = result.get('success', False)
        
        response_payload = {
            "jobId": job_id,
            "status": "success" if is_success else "error",
            "phoneNumber": request.phoneNumber,
            "data": result.get('invoice_data') if is_success else None,
            "error": result.get('error') if not is_success else None
        }
        
        # Send callback if URL provided
        if request.callbackUrl:
            await send_callback(request.callbackUrl, response_payload)
        
        # Return response
        return response_payload
        
    except httpx.HTTPError as e:
        logger.log_error(
            "Failed to download PDF from URL",
            url=request.mediaUrl,
            error=str(e)
        )
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to download PDF: {str(e)}"
        )
        
    except Exception as e:
        logger.log_error(
            "Unexpected error during PDF URL processing",
            url=request.mediaUrl,
            error=str(e)
        )
        
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Internal server error: {str(e)}"
        )


@router.get("/export-excel")
async def export_excel(
    report_type: str = Query("custom", regex="^(daily|monthly|custom)$"),
    start_date: Optional[str] = Query(None, regex="^\d{4}-\d{2}-\d{2}$"),
    end_date: Optional[str] = Query(None, regex="^\d{4}-\d{2}-\d{2}$"),
    phone_number: Optional[str] = Query(None, description="Filter by phone number/company"),
    date_field: str = Query(
        "invoice_date",
        regex="^(invoice_date|created_at)$",
        description="Field to filter by: 'invoice_date' (date on invoice document) or 'created_at' (when saved to database)"
    )
):
    """
    Export invoices and items to Excel and upload to Cloud Storage
    
    - **report_type**: Type of report - 'daily', 'monthly', or 'custom'
    - **start_date**: Start date (YYYY-MM-DD) - required for custom reports
    - **end_date**: End date (YYYY-MM-DD) - required for custom reports
    - **phone_number**: Optional filter by phone number/company
    - **date_field**: Field to filter by - 'invoice_date' (default) or 'created_at'
    
    **Date field options:**
    - `invoice_date` (default): Filter by the date ON the invoice document (for accounting/fiscal reports)
    - `created_at`: Filter by when the invoice was SAVED to database (for operational/audit reports)
    
    Examples:
    - Daily (invoices dated today): `/export-excel?report_type=daily&date_field=invoice_date`
    - Daily (processed today): `/export-excel?report_type=daily&date_field=created_at`
    - Monthly (default): `/export-excel?report_type=monthly&phone_number=51987654321`
    - Monthly (operational): `/export-excel?report_type=monthly&date_field=created_at&phone_number=51987654321`
    - Custom: `/export-excel?report_type=custom&start_date=2024-01-01&end_date=2024-03-31`
    
    Returns JSON with:
    - **file_url**: Public URL to download the Excel file (expires in 1 hour)
    - **filename**: Name of the generated Excel file
    - **statistics**: Invoice and item counts
    - **date_range**: Start and end dates of the report
    
    The Excel file contains 2 sheets:
    1. **Facturas** - Main invoice data
    2. **Items de Facturas** - Invoice line items
    """
    logger.log_info(
        "Excel export requested",
        report_type=report_type,
        start_date=start_date,
        end_date=end_date,
        phone_number=phone_number,
        date_field=date_field
    )
    
    try:
        # Calculate date range based on report type
        # Use local time for better user experience (fixes timezone issues)
        from datetime import timezone
        
        now_local = datetime.now()
        today_local = now_local.date()
        
        logger.log_info(
            "📅 Starting report generation",
            report_type=report_type,
            date_field=date_field,
            today_local=str(today_local),
            now_local=now_local.isoformat(),
            raw_start_date=start_date,
            raw_end_date=end_date,
            phone_number=phone_number,
            filter_explanation=f"Will filter by '{date_field}' field"
        )
        
        if report_type == "daily":
            # Today's invoices using local timezone
            calculated_start = today_local.strftime("%Y-%m-%d")
            calculated_end = today_local.strftime("%Y-%m-%d")
            
            # For timestamp filtering (created_at): use local time converted to UTC for Supabase
            if date_field == "created_at":
                # Start: Today at 00:00:00 local time
                start_local = datetime.combine(today_local, datetime.min.time())
                # End: Current time local
                end_local = now_local
                
                # Convert to UTC for Supabase query
                timestamp_start = start_local.astimezone(timezone.utc).isoformat()
                timestamp_end = end_local.astimezone(timezone.utc).isoformat()
            else:
                timestamp_start = None
                timestamp_end = None
            
            day_name = today_local.strftime('%A')  # Full day name (e.g., "Wednesday")
            report_name = f"Reporte_Diario_{today_local.strftime('%Y%m%d')}"
            
            logger.log_info(
                "📅 DAILY REPORT - Date range calculated",
                current_day=day_name,
                date_local=calculated_start,
                timestamp_start=timestamp_start,
                timestamp_end=timestamp_end,
                report_name=report_name,
                filter_type=f"{date_field} {'(timestamp local->UTC)' if date_field == 'created_at' else '(date)'}",
                description=f"Today ({day_name} local time) filtering by {date_field}"
            )
            
        elif report_type == "monthly":
            # Current month using local timezone
            month_start = today_local.replace(day=1)
            calculated_start = month_start.strftime("%Y-%m-%d")
            calculated_end = today_local.strftime("%Y-%m-%d")
            
            # For timestamp filtering (created_at): use local time converted to UTC
            if date_field == "created_at":
                # Start: 1st day of month at 00:00:00 local time
                start_local = datetime.combine(month_start, datetime.min.time())
                # End: Current time local
                end_local = now_local
                
                # Convert to UTC for Supabase query
                timestamp_start = start_local.astimezone(timezone.utc).isoformat()
                timestamp_end = end_local.astimezone(timezone.utc).isoformat()
            else:
                timestamp_start = None
                timestamp_end = None
            
            # Report name with month and year
            month_name = today_local.strftime('%B')  # Full month name (e.g., "February", "November")
            report_name = f"Reporte_Mensual_{today_local.strftime('%Y%m')}"
            
            logger.log_info(
                "📅 MONTHLY REPORT - Date range calculated",
                current_month=month_name,
                current_year=today_local.year,
                month_start_date_local=calculated_start,
                today_date_local=calculated_end,
                timestamp_start=timestamp_start,
                timestamp_end=timestamp_end,
                report_name=report_name,
                total_days=(today_local - month_start).days + 1,
                filter_type=f"{date_field} {'(timestamp local->UTC)' if date_field == 'created_at' else '(date)'}",
                description=f"From {month_name} 1st (local time) filtering by {date_field}"
            )
            
        else:  # custom
            if not start_date or not end_date:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="start_date and end_date are required for custom reports"
                )
            
            # Validate date range (max 3 months)
            start_dt = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_dt = datetime.strptime(end_date, "%Y-%m-%d").date()
            
            logger.log_info(
                "📅 CUSTOM REPORT - Validating date range",
                start_date=start_date,
                end_date=end_date,
                start_dt=str(start_dt),
                end_dt=str(end_dt)
            )
            
            if end_dt < start_dt:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="end_date must be greater than or equal to start_date"
                )
            
            # Check 3 month limit
            max_end_date = start_dt + timedelta(days=90)
            days_diff = (end_dt - start_dt).days
            
            logger.log_info(
                "📅 Date range validation",
                days_requested=days_diff,
                max_allowed_days=90,
                max_end_date=str(max_end_date),
                is_valid=end_dt <= max_end_date
            )
            
            if end_dt > max_end_date:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Date range cannot exceed 3 months (90 days)"
                )
            
            calculated_start = start_date
            calculated_end = end_date
            # For custom reports, use invoice_date (no timestamp)
            timestamp_start = None
            timestamp_end = None
            report_name = f"Reporte_Personalizado_{start_date}_a_{end_date}"
            
            logger.log_info(
                "✅ CUSTOM REPORT - Date range validated",
                calculated_start=calculated_start,
                calculated_end=calculated_end,
                total_days=days_diff + 1,
                report_name=report_name,
                filter_type="invoice_date (date only)"
            )
        
        # Get data from Supabase
        logger.log_info(
            "🔍 Querying Supabase for invoices",
            chat_id=phone_number or "all_companies",
            start_date=calculated_start,
            end_date=calculated_end,
            timestamp_start=timestamp_start if date_field == "created_at" else None,
            timestamp_end=timestamp_end if date_field == "created_at" else None,
            report_type=report_type,
            date_field=date_field,
            filter_explanation=f"Filtering by {date_field} field"
        )
        
        invoices, items = await supabase_service.get_invoices_with_items(
            chat_id=phone_number,
            start_date=calculated_start,
            end_date=calculated_end,
            timestamp_start=timestamp_start if date_field == "created_at" else None,
            timestamp_end=timestamp_end if date_field == "created_at" else None,
            date_field=date_field,
            report_type=report_type
        )
        
        logger.log_info(
            "📊 Supabase query completed",
            invoices_count=len(invoices),
            items_count=len(items),
            has_invoices=bool(invoices),
            has_items=bool(items)
        )
        
        # Analyze date distribution of retrieved invoices
        if invoices:
            invoice_dates = [inv.get('invoice_date') for inv in invoices if inv.get('invoice_date')]
            created_ats = [inv.get('created_at') for inv in invoices if inv.get('created_at')]
            
            if invoice_dates:
                min_invoice_date = min(invoice_dates)
                max_invoice_date = max(invoice_dates)
                
                logger.log_info(
                    "📅 INVOICE_DATE field in retrieved data (fecha en la factura)",
                    total_invoices=len(invoices),
                    earliest_invoice_date=min_invoice_date,
                    latest_invoice_date=max_invoice_date,
                    unique_invoice_dates=len(set(invoice_dates)),
                    explanation="invoice_date = fecha que aparece EN la factura"
                )
            
            if created_ats:
                min_created = min(created_ats)
                max_created = max(created_ats)
                
                logger.log_info(
                    "📅 CREATED_AT field in retrieved data (cuándo se guardó en BD)",
                    earliest_created_at=min_created,
                    latest_created_at=max_created,
                    explanation="created_at = cuándo se guardó la factura en la base de datos"
                )
            
            # Show filter vs data comparison
            if date_field == "created_at":
                logger.log_info(
                    "🔍 FILTER COMPARISON (using created_at)",
                    filter_field="created_at",
                    requested_timestamp_start=timestamp_start,
                    requested_timestamp_end=timestamp_end,
                    data_created_range=f"{min_created} to {max_created}" if created_ats else "N/A",
                    use_case="Operational/Audit report - filtered by WHEN invoice was SAVED"
                )
            else:
                logger.log_info(
                    "🔍 FILTER COMPARISON (using invoice_date)",
                    filter_field="invoice_date",
                    requested_start=calculated_start,
                    requested_end=calculated_end,
                    data_invoice_date_range=f"{min_invoice_date} to {max_invoice_date}" if invoice_dates else "N/A",
                    dates_match=min_invoice_date >= calculated_start and max_invoice_date <= calculated_end if invoice_dates else False,
                    use_case="Accounting/Fiscal report - filtered by DATE ON INVOICE"
                )
            
            # Log first 3 invoices for debugging
            if invoice_dates:
                sample_invoices = invoices[:3]
                for idx, inv in enumerate(sample_invoices, 1):
                    logger.log_info(
                        f"📄 Sample invoice #{idx}",
                        id=inv.get('id'),
                        invoice_number=inv.get('invoice_number'),
                        invoice_date=inv.get('invoice_date'),
                        created_at=inv.get('created_at'),
                        company_id=inv.get('company_id'),
                        supplier_name=inv.get('supplier_name'),
                        total=inv.get('total'),
                        note="invoice_date = fecha EN la factura | created_at = cuándo se guardó"
                    )
            else:
                logger.log_warning(
                    "⚠️ No invoice_date found in any invoice",
                    total_invoices=len(invoices)
                )
        else:
            logger.log_warning(
                "⚠️ No invoices found in Supabase",
                requested_start=calculated_start,
                requested_end=calculated_end,
                phone_number=phone_number
            )
        
        if not invoices:
            logger.log_warning(
                "⚠️ No invoices found for export - will generate empty Excel",
                report_type=report_type,
                start_date=calculated_start,
                end_date=calculated_end,
                phone_number=phone_number,
                query_details={
                    "chat_id": phone_number or "all",
                    "date_range": f"{calculated_start} to {calculated_end}"
                }
            )
            # Don't raise error, continue to generate empty Excel with headers only
            invoices = []
            items = []

        # Create Excel file
        wb = Workbook()
        
        # ========================================================================
        # SHEET 1: INVOICES
        # ========================================================================
        ws_invoices = wb.active
        ws_invoices.title = "Facturas"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers for invoices
        invoice_headers = [
            "ID", "Job ID", "Compañía/Teléfono", "Nro. Factura", "Fecha Factura",
            "Proveedor", "RUC Proveedor", "Cliente", "RUC Cliente",
            "CC/Placa", "Tipo Documento", "Tipo Costo", "Subtotal", "IGV", "Total", "Moneda", 
            "Confianza", "Archivo Origen", "URL Origen", "Tipo MIME", "Estado",
            "Fecha Creación"
        ]
        
        # Write headers
        for col, header in enumerate(invoice_headers, 1):
            cell = ws_invoices.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        
        # Write invoice data
        for row_idx, invoice in enumerate(invoices, 2):
            ws_invoices.cell(row=row_idx, column=1, value=invoice.get('id'))
            ws_invoices.cell(row=row_idx, column=2, value=invoice.get('job_id'))
            ws_invoices.cell(row=row_idx, column=3, value=invoice.get('company_id'))
            ws_invoices.cell(row=row_idx, column=4, value=invoice.get('invoice_number'))
            ws_invoices.cell(row=row_idx, column=5, value=invoice.get('invoice_date'))
            ws_invoices.cell(row=row_idx, column=6, value=invoice.get('supplier_name'))
            ws_invoices.cell(row=row_idx, column=7, value=invoice.get('supplier_ruc'))
            ws_invoices.cell(row=row_idx, column=8, value=invoice.get('customer_name'))
            ws_invoices.cell(row=row_idx, column=9, value=invoice.get('customer_ruc'))
            ws_invoices.cell(row=row_idx, column=10, value=invoice.get('cc_or_placa'))
            ws_invoices.cell(row=row_idx, column=11, value=invoice.get('document_type'))
            ws_invoices.cell(row=row_idx, column=12, value=invoice.get('tipo_costo'))
            ws_invoices.cell(row=row_idx, column=13, value=invoice.get('subtotal'))
            ws_invoices.cell(row=row_idx, column=14, value=invoice.get('tax'))
            ws_invoices.cell(row=row_idx, column=15, value=invoice.get('total'))
            ws_invoices.cell(row=row_idx, column=16, value=invoice.get('currency'))
            ws_invoices.cell(row=row_idx, column=17, value=invoice.get('confidence_score'))
            ws_invoices.cell(row=row_idx, column=18, value=invoice.get('source_file'))
            ws_invoices.cell(row=row_idx, column=19, value=invoice.get('source_url'))
            ws_invoices.cell(row=row_idx, column=20, value=invoice.get('mime_type'))
            ws_invoices.cell(row=row_idx, column=21, value=invoice.get('processing_status'))
            ws_invoices.cell(row=row_idx, column=22, value=str(invoice.get('created_at', '')))
            
            # Apply borders to data cells
            for col in range(1, len(invoice_headers) + 1):
                ws_invoices.cell(row=row_idx, column=col).border = border
        
        # Auto-adjust column widths
        column_widths = [8, 18, 15, 15, 12, 25, 12, 25, 12, 15, 15, 15, 12, 12, 12, 8, 10, 20, 25, 15, 10, 20]
        for col, width in enumerate(column_widths, 1):
            ws_invoices.column_dimensions[ws_invoices.cell(row=1, column=col).column_letter].width = width
        
        # ========================================================================
        # SHEET 2: INVOICE ITEMS
        # ========================================================================
        ws_items = wb.create_sheet("Items de Facturas")
        
        # Headers for items
        item_headers = [
            "ID Item", "ID Factura", "Nro. Factura", "Fecha Factura",
            "Proveedor", "RUC Proveedor", "Cliente", "Tipo Documento", 
            "CC/Placa", "Moneda", "Compañía/Teléfono", "Nro. Item",
            "Descripción", "Cantidad", "Unidad", "Precio Unitario",
            "Total Item", "Fecha Creación"
        ]
        
        # Write headers
        for col, header in enumerate(item_headers, 1):
            cell = ws_items.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        
        # Write items data
        # Create invoice lookup dictionary for quick access
        invoice_lookup = {inv['id']: inv for inv in invoices}
        
        for row_idx, item in enumerate(items, 2):
            # Get related invoice data
            invoice = invoice_lookup.get(item.get('invoice_id'), {})
            
            ws_items.cell(row=row_idx, column=1, value=item.get('id'))
            ws_items.cell(row=row_idx, column=2, value=item.get('invoice_id'))
            ws_items.cell(row=row_idx, column=3, value=invoice.get('invoice_number'))
            ws_items.cell(row=row_idx, column=4, value=invoice.get('invoice_date'))
            ws_items.cell(row=row_idx, column=5, value=invoice.get('supplier_name'))
            ws_items.cell(row=row_idx, column=6, value=invoice.get('supplier_ruc'))
            ws_items.cell(row=row_idx, column=7, value=invoice.get('customer_name'))
            ws_items.cell(row=row_idx, column=8, value=invoice.get('document_type'))
            ws_items.cell(row=row_idx, column=9, value=invoice.get('cc_or_placa'))
            ws_items.cell(row=row_idx, column=10, value=invoice.get('currency'))
            ws_items.cell(row=row_idx, column=11, value=item.get('company_id'))
            ws_items.cell(row=row_idx, column=12, value=item.get('item_number'))
            ws_items.cell(row=row_idx, column=13, value=item.get('description'))
            ws_items.cell(row=row_idx, column=14, value=item.get('quantity'))
            ws_items.cell(row=row_idx, column=15, value=item.get('unit'))
            ws_items.cell(row=row_idx, column=16, value=item.get('unit_price'))
            ws_items.cell(row=row_idx, column=17, value=item.get('total_price'))
            ws_items.cell(row=row_idx, column=18, value=str(item.get('created_at', '')))
            
            # Apply borders
            for col in range(1, len(item_headers) + 1):
                ws_items.cell(row=row_idx, column=col).border = border
        
        # Auto-adjust column widths
        item_column_widths = [10, 12, 15, 12, 25, 12, 25, 15, 15, 10, 15, 10, 40, 10, 10, 15, 15, 20]
        for col, width in enumerate(item_column_widths, 1):
            ws_items.column_dimensions[ws_items.cell(row=1, column=col).column_letter].width = width
        
        # ========================================================================
        # SAVE AND UPLOAD TO CLOUD STORAGE
        # ========================================================================
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        logger.log_info(
            "Excel file generated in memory",
            report_type=report_type,
            invoices_count=len(invoices),
            items_count=len(items),
            start_date=calculated_start,
            end_date=calculated_end
        )
        
        # Upload to Cloud Storage using storage service
        filename = f"{report_name}.xlsx"
        
        upload_metadata = {
            "report_type": report_type,
            "phone_number": phone_number or "all",
            "start_date": calculated_start,
            "end_date": calculated_end,
            "total_invoices": str(len(invoices)),
            "total_items": str(len(items)),
            "date_field": date_field
        }
        
        upload_result = storage_service.upload_report_excel(
            excel_bytes=excel_file,
            filename=filename,
            phone_number=phone_number,
            metadata=upload_metadata
        )
        
        if not upload_result:
            logger.log_error(
                "Failed to upload Excel to Cloud Storage",
                report_type=report_type
            )
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to upload report to storage"
            )
        
        logger.log_info(
            "✅ Excel uploaded to Cloud Storage",
            blob_path=upload_result["blob_path"],
            url_type=upload_result["url_type"],
            access=upload_result["access"],
            expires_in=upload_result["expires_in"],
            is_empty=len(invoices) == 0
        )
        
        # Determine if report is empty
        is_empty_report = len(invoices) == 0
        
        # Return JSON response with file URL
        return JSONResponse(
            content={
                "file_url": upload_result["file_url"],
                "filename": filename,
                "blob_path": upload_result["blob_path"],
                "report_type": report_type,
                "date_range": {
                    "start": calculated_start,
                    "end": calculated_end
                },
                "statistics": {
                    "total_invoices": len(invoices),
                    "total_items": len(items)
                },
                "is_empty": is_empty_report,
                "message": "No invoices found for the selected date range" if is_empty_report else "Report generated successfully",
                "generated_at": upload_result["generated_at"],
                "url_type": upload_result["url_type"],
                "expires_in": upload_result["expires_in"],
                "access": upload_result["access"]
            },
            headers={
                "X-Total-Invoices": str(len(invoices)),
                "X-Total-Items": str(len(items)),
                "X-Report-Type": report_type,
                "X-Date-Range": f"{calculated_start} to {calculated_end}",
                "X-URL-Type": upload_result["url_type"],
                "X-Is-Empty": "true" if is_empty_report else "false"
            }
        )
        
    except HTTPException:
        raise
        
    except Exception as e:
        logger.log_error(
            "Unexpected error during Excel report generation",
            error=str(e),
            error_type=type(e).__name__,
            report_type=report_type
        )
        import traceback
        logger.log_error(
            "Full traceback",
            traceback=traceback.format_exc()
        )
        
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate report: {str(e)}"
        )


@router.get("/health")
async def health_check():
    """Health check endpoint"""
    return {"status": "healthy", "service": "invoice_extraction"}


@router.get("/info")
async def service_info():
    """Get service information"""
    from config import settings
    
    return {
        "service_name": "WhatsApp Invoice Extraction API",
        "supported_file_types": settings.allowed_extensions,
        "max_file_size_mb": settings.max_file_size_mb,
        "gemini_model": settings.gemini_model,
        "processing_phases": [
            "Phase 1: ZIP validation and extraction",
            "Phase 2: Chronological file sorting", 
            "Phase 3: File type classification",
            "Phase 4: Content processing by type",
            "Phase 5: Structured data extraction",
            "Phase 6: Pipeline orchestration"
        ],
        "single_file_endpoints": [
            "POST /process-image - Process a single image file (multipart/form-data)",
            "POST /process-image-url - Process an image from URL (WhatsApp API)",
            "POST /process-pdf - Process a single PDF file (multipart/form-data)",
            "POST /process-pdf-url - Process a PDF from URL (WhatsApp API)"
        ]
    }

